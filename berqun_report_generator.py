"""
Berqun Rapor Oluşturucu — Aylık verimlilik raporları.
Berqun API'den veri çekip Excel ve HTML rapor oluşturur.
Takımları GMY bazında gruplar.
"""

import os
import json
from datetime import datetime, timedelta
from typing import Optional
from berqun_client import get_berqun_client, PERIOD_MONTH, PERIOD_DAY

# ============================================================
# ORGANİZASYON ŞEMASI — Domain → Tribe Eşleştirmesi
# Berqun'daki team_name → Domain gruplama
# ============================================================

# Berqun team_name prefix → Domain exact mapping
# team_name formatı: "Tribe Adı - Alt Takım - Lider" şeklinde
# Bu map'te en uzun prefix önce kontrol edilir (longest match wins)
TEAM_DOMAIN_MAP = {
    # Wallet & Loyalty
    "Wallet": "Wallet & Loyalty",
    "Loyalty": "Wallet & Loyalty",
    "B2B Wallet Core": "Wallet & Loyalty",
    "B2B2C Proje Uyarlama": "Wallet & Loyalty",
    # Payment Facilitator & Android POS & Soft POS
    "Android POS": "Payment Facilitator & Android POS & Soft POS",
    "SoftPos": "Payment Facilitator & Android POS & Soft POS",
    "Soft POS": "Payment Facilitator & Android POS & Soft POS",
    "PF & POS Product Leads": "Payment Facilitator & Android POS & Soft POS",
    "Payment Facilitator": "Payment Facilitator & Android POS & Soft POS",
    "Mobil": "Payment Facilitator & Android POS & Soft POS",
    "Transit": "Payment Facilitator & Android POS & Soft POS",
    # Zubizu
    "Zubizu": "Zubizu",
    # Kartlı Ödeme Sistemleri
    "Kartlı Ödeme Sistemleri": "Kartlı Ödeme Sistemleri",
    "Fraud, Kampanya, Takas, Kart Basım": "Kartlı Ödeme Sistemleri",
    # Temel Bankacılık
    "Temel Bankacılık": "Temel Bankacılık",
    # Engagement Management & İş Geliştirme ve Satış
    "Engagement Management & İş Geliştirme ve Satış": "Engagement Management & İş Geliştirme ve Satış",
    "Engagement Management": "Engagement Management & İş Geliştirme ve Satış",
    "Yurtdışı Satış": "Engagement Management & İş Geliştirme ve Satış",
    "BT Satın Alma ve Sözleşme Yönetimi": "Engagement Management & İş Geliştirme ve Satış",
    "Sözleşme ve Kaynak Yönetimi": "Engagement Management & İş Geliştirme ve Satış",
    # B2B2C
    "B2B2C": "B2B2C",
    # Operasyon
    "Operasyon": "Operasyon",
    # IT Altyapı ve Platform Teknolojileri (Bilgi Güvenliği dahil)
    "Altyapı, Sistem ve Operasyon": "IT Altyapı ve Platform Teknolojileri",
    "Altyapı, Sistem Operasyon ve Güvenlik": "IT Altyapı ve Platform Teknolojileri",
    "Altyapı,Sistem ve Operasyon": "IT Altyapı ve Platform Teknolojileri",
    "IT Altyapı ve Platform Teknolojileri": "IT Altyapı ve Platform Teknolojileri",
    "IT Altyapı": "IT Altyapı ve Platform Teknolojileri",
    "Uygulama Yönetimi": "IT Altyapı ve Platform Teknolojileri",
    "Teknik İşletim Planlama": "IT Altyapı ve Platform Teknolojileri",
    "Teknik İşletim &": "IT Altyapı ve Platform Teknolojileri",
    "Teknik İşletim": "IT Altyapı ve Platform Teknolojileri",
    "Bilgi Güvenliği ve İş Sürekliliği": "IT Altyapı ve Platform Teknolojileri",
    "Bilgi Güvenliği": "IT Altyapı ve Platform Teknolojileri",
    "DevOps": "IT Altyapı ve Platform Teknolojileri",
    "Veri Tabanı": "IT Altyapı ve Platform Teknolojileri",
    # Strateji, Dijitalleşme ve Portföy Yönetimi
    "Strateji, Dijitalleşme ve Portföy Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Strateji ve Dijitalleşme": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Kurumsal Destek ve Hizmet Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Portföy ve Süreç Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Süreç ve Ölçme Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Talep ve Portföy Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Test Otomasyon ve Kalite Araçları Yönetimi": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Deneyim Tasarımı": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Dijitalleşme ve RPA": "Strateji, Dijitalleşme ve Portföy Yönetimi",
}

# Kişi ad-soyad → domain eşleştirmesi (Berqun'da team_name bulunamazsa fallback)
PERSON_DOMAIN_MAP = {
    "Belkay Sarataş": "Temel Bankacılık",
    "Nisa Nur Gürsoy": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Akın Uğur Özcan": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "İrem Nur Güneş": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Osman Bozok": "Temel Bankacılık",
    "Sude Yalçınkaya": "Temel Bankacılık",
    "Bayram Yıldırım": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Taşkın Yüzel": "IT Altyapı ve Platform Teknolojileri",
    "İlhan Yiğit": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Can Kadıoğlu": "IT Altyapı ve Platform Teknolojileri",
    "Cansu Ulusoy": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Gökhan Leblebici": "Kartlı Ödeme Sistemleri",
    "Berat Cem Özçelik": "Temel Bankacılık",
    "Fatma Bilgin": "Temel Bankacılık",
    "Seçil Şen Dalda": "Kartlı Ödeme Sistemleri",
    "Nilsu Acar": "Temel Bankacılık",
    "Eyüp Şeşen": "Wallet & Loyalty",
    "Ahmet Mesut Erçıkan": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Bekir Sıddık Adıgüzel": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Cihan Topçu": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Okan Can Karadağ": "Temel Bankacılık",
    "Zeki Gülen": "Payment Facilitator & Android POS & Soft POS",
    "İbrahim Tarhan": "Temel Bankacılık",
    "Abbas Ayaz": "Temel Bankacılık",
    "Suhat Baykal": "IT Altyapı ve Platform Teknolojileri",
    "Fatih Ergül": "Payment Facilitator & Android POS & Soft POS",
    "Mehmet Furkan Eser": "Kartlı Ödeme Sistemleri",
    "Mustafa Menteşoğlu": "Payment Facilitator & Android POS & Soft POS",
    "Eren Atbaş": "Payment Facilitator & Android POS & Soft POS",
    "Merve Esen": "Temel Bankacılık",
    "Zafer Özdemir": "IT Altyapı ve Platform Teknolojileri",
    "Alp Rıza Ünal": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Betül Şinik": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Erdem Bilir": "IT Altyapı ve Platform Teknolojileri",
    "Fatih Sultan Mehmet Yiğit": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Gamze Demirsoy": "Kartlı Ödeme Sistemleri",
    "Melih Alkan": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Tuncay Uyar": "Payment Facilitator & Android POS & Soft POS",
    "Abdülmecit Ekinci": "Payment Facilitator & Android POS & Soft POS",
    "Arif Emre Kütük": "Kartlı Ödeme Sistemleri",
    "Gürel Yıldız": "Payment Facilitator & Android POS & Soft POS",
    "Yusuf Yıldırım": "Kartlı Ödeme Sistemleri",
    "Caner Abalı": "Payment Facilitator & Android POS & Soft POS",
    "Evrim Büyükduman": "Kartlı Ödeme Sistemleri",
    "Abdullah Karataş": "Temel Bankacılık",
    "Caner Akın": "Temel Bankacılık",
    "Hasan Özgür Güçlü": "Kartlı Ödeme Sistemleri",
    "Serhat Tarhan": "Kartlı Ödeme Sistemleri",
    "Emrah Gündüz": "Wallet & Loyalty",
    "Can Alpay Çiftçi": "Kartlı Ödeme Sistemleri",
    "Celalettin Aksoy": "IT Altyapı ve Platform Teknolojileri",
    "Ethem Boynukara": "Temel Bankacılık",
    "Ramazan Aksoy": "Kartlı Ödeme Sistemleri",
    "Çağla Başak": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Emre Ermiş": "IT Altyapı ve Platform Teknolojileri",
    "Gökhan Terzioğlu": "Payment Facilitator & Android POS & Soft POS",
    "Hakan Kaysi": "Temel Bankacılık",
    "Kadir İlyasoğlu": "Kartlı Ödeme Sistemleri",
    "Kübra Kaçar": "Kartlı Ödeme Sistemleri",
    "Süleyman Koparır": "Wallet & Loyalty",
    "Alp Eren Bal": "Payment Facilitator & Android POS & Soft POS",
    "Arda Tunçalp": "Kartlı Ödeme Sistemleri",
    "Cemaleddin Kuru": "Temel Bankacılık",
    "Rıdvan Çakmak": "Payment Facilitator & Android POS & Soft POS",
    "Ümit Furkan": "Kartlı Ödeme Sistemleri",
    "Emine Demirşen": "Wallet & Loyalty",
    "Fatma Betül Demirci": "Wallet & Loyalty",
    "Barış Gökşen": "Kartlı Ödeme Sistemleri",
    "Damla Acun": "Temel Bankacılık",
    "Göktürk Erdoğan": "Kartlı Ödeme Sistemleri",
    "Hüsnü Akçak": "Temel Bankacılık",
    "İlker Yavuz": "Kartlı Ödeme Sistemleri",
    "Muhammed Can Gültekin": "Wallet & Loyalty",
    "Ali Gönen": "IT Altyapı ve Platform Teknolojileri",
    "Alperen Erdoğan": "Wallet & Loyalty",
    "Aygün Terzi": "IT Altyapı ve Platform Teknolojileri",
    "Şamil Kaya": "Kartlı Ödeme Sistemleri",
    "Yunus Görüm": "IT Altyapı ve Platform Teknolojileri",
    "Büşra Kaba": "Payment Facilitator & Android POS & Soft POS",
    "Emre Savaştürk": "IT Altyapı ve Platform Teknolojileri",
    "Aykut Eryol": "Payment Facilitator & Android POS & Soft POS",
    "Mert Orhan Güncer": "Payment Facilitator & Android POS & Soft POS",
    "Özge Akyürek": "Temel Bankacılık",
    "Tuba Sezer Şahin": "Payment Facilitator & Android POS & Soft POS",
    "İrem Uluç": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Nesrin Arslan": "Payment Facilitator & Android POS & Soft POS",
    "Asüde Aydın": "Kartlı Ödeme Sistemleri",
    "Bilal Özgül": "Payment Facilitator & Android POS & Soft POS",
    "Hüseyin Doğru": "Temel Bankacılık",
    "Muhammed Taşbaşı": "IT Altyapı ve Platform Teknolojileri",
    "Oğuzhan Akyol": "IT Altyapı ve Platform Teknolojileri",
    "Zafer Kasap": "Temel Bankacılık",
    "Berkant Şencer": "Kartlı Ödeme Sistemleri",
    "Bilal Orkan Tanır": "Temel Bankacılık",
    "Halis Elüstü": "IT Altyapı ve Platform Teknolojileri",
    "Cüneyt Eren": "Wallet & Loyalty",
    "Erdinç Kaçar": "Kartlı Ödeme Sistemleri",
    "Göksel Ateş": "Wallet & Loyalty",
    "Hüseyin Emre Çevik": "Temel Bankacılık",
    "Mehmet Şükrü Kavak": "Payment Facilitator & Android POS & Soft POS",
    "Volkan Sarsan": "Kartlı Ödeme Sistemleri",
    "Zeynep Demiray": "Wallet & Loyalty",
    "Caner Çelik": "IT Altyapı ve Platform Teknolojileri",
    "İbrahim Demir": "Payment Facilitator & Android POS & Soft POS",
    "Merve Akçidem": "Kartlı Ödeme Sistemleri",
    "Merve Aktaş": "Temel Bankacılık",
    "Muhammed Turan": "Temel Bankacılık",
    "Myriem Belkacemi": "Temel Bankacılık",
    "Neşe Abbak": "Temel Bankacılık",
    "Özlem Küçük": "Temel Bankacılık",
    "Cavit Sevinç": "Wallet & Loyalty",
    "Cem Bozlar": "Kartlı Ödeme Sistemleri",
    "Ömer Öztürk": "IT Altyapı ve Platform Teknolojileri",
    "İclal Aslan": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Zekeriyya Özkul": "Kartlı Ödeme Sistemleri",
    "Mustafa Yıldırım": "Temel Bankacılık",
    "Naim Yıldız": "Temel Bankacılık",
    "Kemal Genç": "Temel Bankacılık",
    "Mesut Işık": "Wallet & Loyalty",
    "Murat Ergün": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Yasin Çetin": "Wallet & Loyalty",
    "Canan Korkmaz Aslıtürk": "IT Altyapı ve Platform Teknolojileri",
    "Mustafa Bülbül": "IT Altyapı ve Platform Teknolojileri",
    "Emre Özgöz": "Temel Bankacılık",
    "Merve Amet": "Payment Facilitator & Android POS & Soft POS",
    "Ahmet Pirimoğlu": "Temel Bankacılık",
    "Alper Şık": "IT Altyapı ve Platform Teknolojileri",
    "Bilgehan Takım": "Temel Bankacılık",
    "Cebrail Karataş": "Temel Bankacılık",
    "Çağlar Çakır": "Kartlı Ödeme Sistemleri",
    "Eren Aydın": "IT Altyapı ve Platform Teknolojileri",
    "Levent Çiflik": "Temel Bankacılık",
    "Melek Demirtaş": "Kartlı Ödeme Sistemleri",
    "Mücahit İmre": "Payment Facilitator & Android POS & Soft POS",
    "Mehmet Erol": "Kartlı Ödeme Sistemleri",
    "Zafer Aydınlı": "Payment Facilitator & Android POS & Soft POS",
    "Ali İhsan Basar": "Wallet & Loyalty",
    "Murat Atlıoğlu": "Wallet & Loyalty",
    "Muratcan Tepençelik": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Ercan Tetik": "Payment Facilitator & Android POS & Soft POS",
    "Esra Yenidünya": "Wallet & Loyalty",
    "Revan Ali": "Payment Facilitator & Android POS & Soft POS",
    "Bülent Bingöl": "IT Altyapı ve Platform Teknolojileri",
    "Okan Geler": "IT Altyapı ve Platform Teknolojileri",
    "Burak Keleşoğlu": "IT Altyapı ve Platform Teknolojileri",
    "Elif Didem Durmaz": "IT Altyapı ve Platform Teknolojileri",
    "Serdar Özaydın": "IT Altyapı ve Platform Teknolojileri",
    "Rıdvan Özhan": "Temel Bankacılık",
    "Gülbeyaz Akgüç Babacan": "Kartlı Ödeme Sistemleri",
    "Müjdat Karabulut": "Kartlı Ödeme Sistemleri",
    "Sinan Yılmaz": "Payment Facilitator & Android POS & Soft POS",
    "Necdet Bilik": "IT Altyapı ve Platform Teknolojileri",
    "Semin Cintan Eray": "Kartlı Ödeme Sistemleri",
    "İpek Dereli": "Payment Facilitator & Android POS & Soft POS",
    "Ahmet Hasan Acar": "IT Altyapı ve Platform Teknolojileri",
    "Defne Acar": "Temel Bankacılık",
    "Ömer Hadi Altunkaya": "IT Altyapı ve Platform Teknolojileri",
    "Ali Faruk Gürkan": "IT Altyapı ve Platform Teknolojileri",
    "Aykut Gürsel": "Payment Facilitator & Android POS & Soft POS",
    "Murat Can Arıgün": "Kartlı Ödeme Sistemleri",
    "Serhat İnan": "IT Altyapı ve Platform Teknolojileri",
    "Abdülhamid Bayrak": "Wallet & Loyalty",
    "Emre Çoğalan": "Kartlı Ödeme Sistemleri",
    "Faruk Cihan Özboz": "Payment Facilitator & Android POS & Soft POS",
    "Ahmet Öztürk": "Engagement Management & İş Geliştirme ve Satış",
    "Esin Küçükbaş": "Kartlı Ödeme Sistemleri",
    "Oğuzhan Çengiz": "Kartlı Ödeme Sistemleri",
    "Yasir Emin Çiftçi": "Kartlı Ödeme Sistemleri",
    "Hilmi İltar": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Melike Alpaslan": "Kartlı Ödeme Sistemleri",
    "Omar Gobran": "Payment Facilitator & Android POS & Soft POS",
    "Seçkin Öter": "IT Altyapı ve Platform Teknolojileri",
    "Gökhan Yılmaz": "IT Altyapı ve Platform Teknolojileri",
    "Mehmet Fatih Dayan": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Fatma Nur Karabulut": "Payment Facilitator & Android POS & Soft POS",
    "Semih Akcan": "Temel Bankacılık",
    "Ali İrim": "Kartlı Ödeme Sistemleri",
    "Çağrı Korkmazgöz": "IT Altyapı ve Platform Teknolojileri",
    "Gülcan Çetinli": "Engagement Management & İş Geliştirme ve Satış",
    "Ömer Faruk Bayrak": "IT Altyapı ve Platform Teknolojileri",
    "Ekrem Kürkçü": "Temel Bankacılık",
    "Sabri Demirel": "Temel Bankacılık",
    "Gökmen Çalışan": "Engagement Management & İş Geliştirme ve Satış",
    "Emin Uzun": "Temel Bankacılık",
    "Erhan Akbaş": "Temel Bankacılık",
    "Erkan Çokbekler": "Payment Facilitator & Android POS & Soft POS",
    "Doruk Ertoğlu": "Kartlı Ödeme Sistemleri",
    "Yunus Kandemir": "Temel Bankacılık",
    "Muhammed Özcan": "Payment Facilitator & Android POS & Soft POS",
    "Mahmut Sarıhan": "Payment Facilitator & Android POS & Soft POS",
    "Tayfun Evrim Çakırtaş": "Temel Bankacılık",
    "Mehmet Yılmaz": "Engagement Management & İş Geliştirme ve Satış",
    "Rasih Baysal": "Kartlı Ödeme Sistemleri",
    "Burak Eder": "Engagement Management & İş Geliştirme ve Satış",
    "Kamil Köse": "Engagement Management & İş Geliştirme ve Satış",
    "Esra Hayta": "Engagement Management & İş Geliştirme ve Satış",
    "Umut Minarli": "Strateji, Dijitalleşme ve Portföy Yönetimi",
    "Ayla Özer": "Temel Bankacılık",
    "Baran Çağdaş": "Engagement Management & İş Geliştirme ve Satış",
    "Dilruba Yazıcı": "Wallet & Loyalty",
    "Kübra Kaya Köseoğlu": "IT Altyapı ve Platform Teknolojileri",
    "Merve Kocabin": "Payment Facilitator & Android POS & Soft POS",
    "Emre Karaaslan": "IT Altyapı ve Platform Teknolojileri",
}

# Operasyon (büyük ekip — kısa tutuluyor)
_OPERASYON_NAMES = [
    "Nurgül Akosman", "Burak Ömer Şahin", "Nizamettin Özbek", "Abdulsamet Coşkuner",
    "Serhat Taşan", "Faruk Salmış", "Evren Gülcemal", "Melis Işıkyıldız Yalçınkaya",
    "Münevver Akbaş", "Nagihan Varol", "Ersan Balata", "Seda Bölge", "Semih Temel",
    "Ersin Biçer", "Muhammed Kaya", "Samet Yıldırım", "Abdülkadir Altınbay",
    "Ziya Onur Yavuz", "Yeşim Karakaya", "Serap Akın", "Birgül Vit",
    "Berker Girgin", "Feyza Peşte", "Sübeyla Bayındır", "Osman Karakoç",
    "Sercan Sayhan", "Duygu Kaya Dugan", "Gülbahar Eken Çalışçı", "Tolga Demirer",
    "Ümit Karasakal", "Çağla Aydın", "Aylin Sümer", "Berat Kılıç", "Mert Güler",
    "Ozan Okçu", "Yazgı Gülcivan", "Engin Gökgöz", "İpek Çetin", "Serap Çapkan",
    "Nihal Temel", "Sinem Seyrek", "Tuğba Sarı", "Cüneyt Yıldırım",
    "Osmancan Paray", "Zerin Ongun", "Ayşegül Kaleli", "Hüseyin Ali Ünlü",
    "Sevginar Korkmaz", "Yasemin Soysal", "Çiğdem Aslan", "Umur Hüseyinoğlu",
    "Didem Şahin", "Sare Gümüş", "Özge Kalaycıoğlu", "Mert Kılıç", "Funda Çabuş",
    "Burak Ocak", "Kadriye Ceylan", "Ümit Durusal", "Yılmaz Kayalar",
    "Alper Aydın", "Özlem Harman", "Şükrü Özçelik", "Gülsüm Kosif",
    "Meryem Kaya", "Batuhan Oluk", "Saadet Canatan", "Beyhan Güngör", "Ebru Can",
    "Hande Selin Er", "Ahmet Refik Gülenler", "Gizem Elif Şahin", "Zeynep Bulut",
    "Erdem Ekşi", "Serap Altunok", "Haluk Yılmaz", "Burhan Tepe",
    "Anıl Can Taşkıran", "Handan Önem", "Sinan Ünver", "İrem Beyzi",
    "Deste Çelik", "Engin Sağ", "Gizem Emre"
]
for _name in _OPERASYON_NAMES:
    PERSON_DOMAIN_MAP[_name] = "Operasyon"

# Zubizu
_ZUBIZU_NAMES = [
    "Anıl Sağlam", "Bilge Güven", "Cemre Ekşioğlu", "Cenk Ayyıldız",
    "Derya Ateş", "Dilek Haydaroğulları", "Ebru Sunter", "Gizem Sirimsi",
    "Hicret Şen", "İpek Tağkuş Avkan", "Merve Aydoğan Akbaba", "Merve Erten",
    "Merve Sözen", "Nur Başkan Topuz", "Ömer Aydın", "Pınar Ekinci",
    "Raif Telli", "Sertaç Taşkın", "Tuğçe Torun", "Zeynep Yüksel"
]
for _name in _ZUBIZU_NAMES:
    PERSON_DOMAIN_MAP[_name] = "Zubizu"

# B2B2C
_B2B2C_NAMES = [
    "Barış Atalay", "Barış Şirin", "Behiye Evren", "Benan Ciuca",
    "Bilal Bedri Balcioglu", "Cihan Güneş", "Esra Kapıyoldaş", "Fatma Kalkan",
    "Görkem Memi", "Kadir Duran", "Teoman Sarıgül", "Yusuf Karahan"
]
for _name in _B2B2C_NAMES:
    PERSON_DOMAIN_MAP[_name] = "B2B2C"


def classify_team_to_domain(team_name: str, staff_name: str = None) -> str:
    """
    Berqun team_name'ini gerçek Domain'e eşleştir.

    Strateji:
    1. Önce kişi adını PERSON_DOMAIN_MAP'te ara (kesin eşleşme)
    2. Sonra team_name'i TEAM_DOMAIN_MAP'teki prefix'lerle karşılaştır
       (en uzun eşleşen prefix kazanır)
    3. Bulunamazsa "Diğer" döndür
    """
    # 1. Kişi adı ile kesin eşleşme
    if staff_name:
        for person, domain in PERSON_DOMAIN_MAP.items():
            if _name_match(staff_name, person):
                return domain

    # 2. Team name prefix match (longest match wins)
    if team_name:
        best_match = None
        best_len = 0
        for prefix, domain in TEAM_DOMAIN_MAP.items():
            if team_name.startswith(prefix) and len(prefix) > best_len:
                best_match = domain
                best_len = len(prefix)
        if best_match:
            return best_match

    return "Diğer"


def _name_match(berqun_name: str, org_name: str) -> bool:
    """İki ismi karşılaştır (Berqun vs Org şeması)"""
    if not berqun_name or not org_name:
        return False
    # Tam eşleşme
    if berqun_name.strip().lower() == org_name.strip().lower():
        return True
    # Berqun'da ad ve soyad ayrı gelebilir
    berqun_parts = set(berqun_name.strip().lower().split())
    org_parts = set(org_name.strip().lower().split())
    # En az 2 kelime eşleşmeli (ad + soyad)
    common = berqun_parts & org_parts
    return len(common) >= 2


# Eski API uyumluluğu için alias
def classify_team_to_gmy(team_name: str, staff_name: str = None) -> str:
    """classify_team_to_domain için uyumluluk alias'ı"""
    return classify_team_to_domain(team_name, staff_name)


def group_staff_by_team(staff_list: list) -> dict:
    """Personeli team_name bazında grupla"""
    teams = {}
    for staff in staff_list:
        team = staff.get("team_name", "Bilinmeyen Takım")
        if team not in teams:
            teams[team] = []
        teams[team].append(staff)
    return teams


def group_staff_by_gmy(staff_list: list) -> dict:
    """Personeli Domain (GMY) bazında grupla"""
    gmy_groups = {}
    for staff in staff_list:
        team = staff.get("team_name", "")
        staff_name = staff.get('full_name', f"{staff.get('name', '')} {staff.get('surname', '')}").strip()
        domain = classify_team_to_domain(team, staff_name)
        if domain not in gmy_groups:
            gmy_groups[domain] = {"teams": {}, "total_staff": 0}
        team_key = team or "Bilinmeyen"
        if team_key not in gmy_groups[domain]["teams"]:
            gmy_groups[domain]["teams"][team_key] = []
        gmy_groups[domain]["teams"][team_key].append(staff)
        gmy_groups[domain]["total_staff"] += 1
    return gmy_groups


# ============================================================
# RAPOR OLUŞTURMA
# ============================================================

def _format_time(val):
    """'10:20:16.461538' → '10:20' formatına çevir."""
    if not val or not isinstance(val, str):
        return val
    parts = val.split(":")
    return f"{parts[0]}:{parts[1]}" if len(parts) >= 2 else val

def generate_team_summary_html(team_name: str, staff_data: list, report_date: str) -> str:
    """
    Takım özet tablosu — Excel Resim 2 formatında HTML tablo.
    Kolonlar: Ad Soyad, Takım, Beklenen Gün, Çalışılan Gün, BQ(%),
    Beklenen Mesai, Gerçekleşen Mesai, Verimli, Verimsiz, İnternet (GB)
    """
    rows = ""
    for staff in staff_data:
        dash = staff.get("dashboard", {})
        name = staff.get("full_name", f"{staff.get('name', '')} {staff.get('surname', '')}")
        
        # Mesaili BQ varsa onu kullan
        mesaili_bq = staff.get("mesaili_bq")
        bq = mesaili_bq if mesaili_bq is not None else _safe_get(dash, "user_bq_point_avg", "bq_score", "bq")
        
        expected_days = _calc_work_days(dash)
        worked_days = _calc_worked_days(dash)
        expected_time = _seconds_to_hhmm(_safe_get(dash, "expected_work_duration_avg"))
        actual_time = _seconds_to_hhmm(_safe_get(dash, "work_duration_avg"))
        productive = _seconds_to_hhmm(_safe_get(dash, "productive_duration_avg"))
        unproductive = _seconds_to_hhmm(_safe_get(dash, "distractive_duration_avg"))
        idle = _seconds_to_hhmm(_safe_get(dash, "idle_duration_avg"))
        internet = _bytes_to_gb(_safe_get(dash, "all_download_bytes", "download_total"))
        first_act = _safe_get(dash, "all_start_time", "start_time_avg", "in_start_time")
        last_act = _safe_get(dash, "all_end_time", "end_time_avg", "in_end_time")

        # BQ puana göre renk
        bq_val = _to_float(bq)
        bq_color = "#e74c3c" if bq_val and bq_val < 100 else "#27ae60" if bq_val else "#7f8c8d"

        rows += f"""
        <tr>
            <td style="font-weight:600;">{name}</td>
            <td>{staff.get('team_name', '-')}</td>
            <td style="text-align:center;">{_fmt(expected_days)}</td>
            <td style="text-align:center;">{_fmt(worked_days)}</td>
            <td style="text-align:center; color:{bq_color}; font-weight:700;">{_fmt(bq)}</td>
            <td style="text-align:center;">{_fmt(expected_time)}</td>
            <td style="text-align:center;">{_fmt(actual_time)}</td>
            <td style="text-align:center;">{_fmt(first_act)}</td>
            <td style="text-align:center;">{_fmt(last_act)}</td>
            <td style="text-align:center; color:#27ae60;">{_fmt(productive)}</td>
            <td style="text-align:center; color:#e74c3c;">{_fmt(unproductive)}</td>
            <td style="text-align:center; color:#7f8c8d;">{_fmt(idle)}</td>
            <td style="text-align:center;">{_fmt(internet)}</td>
        </tr>"""

    return f"""
    <h3 style="color:#1a5276; border-bottom:2px solid #2980b9; padding-bottom:8px;">
        📊 {team_name} — Takım Özet Raporu ({report_date})
    </h3>
    <table style="width:100%; border-collapse:collapse; font-size:12px; font-family:Calibri, Arial;">
        <thead>
            <tr style="background:#1a5276; color:white;">
                <th style="padding:8px; text-align:left;">Ad Soyad</th>
                <th style="padding:8px; text-align:left;">Takım</th>
                <th style="padding:8px;">Beklenen Gün</th>
                <th style="padding:8px;">Çalışılan Gün</th>
                <th style="padding:8px;">BQ (%)</th>
                <th style="padding:8px;">Beklenen Mesai</th>
                <th style="padding:8px;">Gerçekleşen Mesai</th>
                <th style="padding:8px;">İlk Aktivite</th>
                <th style="padding:8px;">Son Aktivite</th>
                <th style="padding:8px;">Verimli</th>
                <th style="padding:8px;">Verimsiz</th>
                <th style="padding:8px;">Aktivite Yok</th>
                <th style="padding:8px;">İnternet (GB)</th>
            </tr>
        </thead>
        <tbody>{rows}
        </tbody>
    </table>"""


def generate_mail_html(gmy_name: str, report_date: str,
                       teams_data: dict, bq_below_100: list,
                       vpn_missing: list, risk_text: str = None) -> str:
    """
    Tam mail HTML'i oluştur — GMY bazında.

    Args:
        gmy_name: GMY adı
        report_date: Rapor dönemi (ör: Ocak 2026)
        teams_data: {team_name: [staff_with_dashboard_data]}
        bq_below_100: BQ < 100 olan kişi listesi
        vpn_missing: 3+ gün VPN yapmayan kişi listesi
        risk_text: Riskli davranış açıklaması
    """
    risk_section = risk_text or "Riskli durum izlenmemiştir."

    # Takım tabloları
    team_tables = ""
    for team_name, staff_list in teams_data.items():
        team_tables += generate_team_summary_html(team_name, staff_list, report_date)
        team_tables += "<br/>"

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Calibri, Arial, sans-serif; color: #2c3e50; line-height: 1.6; }}
            table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
            th, td {{ border: 1px solid #bdc3c7; padding: 6px 10px; }}
            tr:nth-child(even) {{ background-color: #f8f9fa; }}
            tr:hover {{ background-color: #eaf2f8; }}
            .header {{ background: linear-gradient(135deg, #1a5276, #2980b9); color: white; padding: 20px; border-radius: 8px; }}
            .stat-box {{ display: inline-block; background: #f0f3f4; border-radius: 8px; padding: 15px 25px; margin: 5px; text-align: center; }}
            .stat-number {{ font-size: 28px; font-weight: 700; }}
            .warning {{ color: #e74c3c; }}
            .success {{ color: #27ae60; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1 style="margin:0;">📊 Verimlilik Analizi Raporu</h1>
            <p style="margin:5px 0 0 0; opacity:0.9;">{gmy_name} — {report_date}</p>
        </div>

        <br/>
        <p>Merhaba,</p>
        <p>Verimlilik Analizi uygulamasında yapılan <strong>Çalışan Verimliliği</strong> kapsamında,
        çalışanlarımızın verimlilik raporları GMY'lik bazında düzenlenmiştir.</p>

        <p>Verimlilik Analizi Raporlar, çalışanlarımızın gelişimini desteklemek ve şirketimizin
        hedeflerine daha etkin bir şekilde ulaşmasına katkı sunmak amacıyla paylaşılmaktadır.</p>

        <p>Raporların içeriği, çalışanlarımızın günlük faaliyetleri, verimlilik göstergeleri ve
        belirlenen BQ puanı hedefine ulaşma durumu gibi önemli verileri kapsamaktadır.</p>

        <!-- Özet İstatistikler -->
        <div style="margin:20px 0;">
            <div class="stat-box">
                <div class="stat-number warning">{len(bq_below_100)}</div>
                <div>BQ &lt; 100 Kişi</div>
            </div>
            <div class="stat-box">
                <div class="stat-number warning">{len(vpn_missing)}</div>
                <div>3+ Gün VPN Yok</div>
            </div>
        </div>

        <p><strong>Domain özelinde BQ puanı 100 altında olan kişi sayısı:</strong>
        <span class="warning" style="font-size:18px; font-weight:700;">{len(bq_below_100)}</span></p>

        <p><strong>3 gün ve üzeri VPN yapmayan kişi sayısı:</strong>
        <span class="warning" style="font-size:18px; font-weight:700;">{len(vpn_missing)}</span></p>

        <p><strong>Riskli Davranış Örüntüleri Değerlendirmesi:</strong><br/>
        {risk_section}</p>

        <hr style="border:1px solid #eee; margin:20px 0;"/>

        <!-- BQ < 100 Tablosu -->
        {"" if not bq_below_100 else _bq_below_100_table(bq_below_100)}

        <!-- VPN Tablosu -->
        {"" if not vpn_missing else _vpn_missing_table(vpn_missing)}

        <hr style="border:1px solid #eee; margin:20px 0;"/>

        <!-- Takım Tabloları -->
        {team_tables}

        <hr style="border:1px solid #eee; margin:20px 0;"/>

        <h3>📋 Rapor Türleri</h3>
        <ul>
            <li><strong>Çalışan Verimliliği:</strong> Kişi ve takım bazında BQ puanı, verimli/verimsiz
            süreler, fazla mesai, hareketsiz süreler, işe başlama/bitirme süreleri.</li>
            <li><strong>100 Puan Altı BQ:</strong> BQ &lt; 100 personelin detaylı faaliyet verileri.</li>
            <li><strong>3+ Gün VPN Yapmayanlar:</strong> VPN bağlantısı olmayan personel listesi.</li>
        </ul>

        <p>Verimlilik Analizi raporunda, personelin kullanmış olduğu Verimsiz uygulamaların
        Verimli olduğunu düşünüyorsanız, bu uygulamaları
        <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a>
        mail adresine iletebilirsiniz.</p>

        <p style="color:#7f8c8d; font-size:12px;">
        Raporlama aylık periyotlarda yapılacaktır.<br/>
        Destek ve ihtiyaçlarınız için:
        <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a>
        </p>

        <p>Bu süreçte göstereceğiniz anlayış ve iş birliği için şimdiden teşekkür ederiz.</p>
        <p><strong>Saygılarımızla,</strong></p>
    </body>
    </html>"""


def _bq_below_100_table(bq_list: list) -> str:
    """BQ < 100 personel tablosu"""
    rows = ""
    for p in sorted(bq_list, key=lambda x: x.get("bq_score", 0)):
        bq = p.get("bq_score", "-")
        rows += f"""
        <tr>
            <td>{p.get('name', '-')}</td>
            <td>{p.get('team_name', '-')}</td>
            <td style="text-align:center; color:#e74c3c; font-weight:700;">{bq}</td>
            <td>{p.get('email', '-')}</td>
        </tr>"""
    return f"""
    <h3 style="color:#e74c3c;">⚠️ BQ Puanı 100 Altında Olan Personel</h3>
    <table>
        <thead>
            <tr style="background:#e74c3c; color:white;">
                <th style="padding:8px;">Ad Soyad</th>
                <th style="padding:8px;">Takım</th>
                <th style="padding:8px;">BQ Puanı</th>
                <th style="padding:8px;">E-Posta</th>
            </tr>
        </thead>
        <tbody>{rows}</tbody>
    </table><br/>"""


def _vpn_missing_table(vpn_list: list) -> str:
    """VPN 3+ gün yapmayanlar tablosu"""
    rows = ""
    for p in vpn_list:
        rows += f"""
        <tr>
            <td>{p.get('name', '-')}</td>
            <td>{p.get('team_name', '-')}</td>
            <td>{p.get('email', '-')}</td>
            <td style="color:#e67e22;">{p.get('last_activity', '-')}</td>
        </tr>"""
    return f"""
    <h3 style="color:#e67e22;">🔌 3 Gün ve Üzeri VPN Yapmayanlar</h3>
    <table>
        <thead>
            <tr style="background:#e67e22; color:white;">
                <th style="padding:8px;">Ad Soyad</th>
                <th style="padding:8px;">Takım</th>
                <th style="padding:8px;">E-Posta</th>
                <th style="padding:8px;">Son Aktivite</th>
            </tr>
        </thead>
        <tbody>{rows}</tbody>
    </table><br/>"""


# ============================================================
# EXCEL RAPOR OLUŞTURMA (openpyxl)
# ============================================================

def generate_team_excel(team_name: str, staff_data: list, report_date: str, output_dir: str) -> str:
    """Takım bazlı Excel raporu oluştur"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️ openpyxl yüklü değil. pip install openpyxl")
        return ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Takım Raporu"

    # Stiller
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    bq_low_font = Font(name="Calibri", size=10, color="E74C3C", bold=True)
    bq_ok_font = Font(name="Calibri", size=10, color="27AE60", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Başlık satırı
    headers = [
        "Ad Soyad", "Takım Adı", "Beklenen Gün Sayısı", "Çalışılan Gün",
        "BQ (%)",
        "İlk Aktivite", "Son Aktivite", "Beklenen Verimli", "Verimli (%)",
        "Verimsiz", "Verimli+Verimsiz", "Toplam Verimli(%)",
        "Toplam Aktivite Yok", "İnternet (GB)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Veri satırları
    for row_idx, staff in enumerate(staff_data, 2):
        dash = staff.get("dashboard", {})
        name = staff.get("full_name", f"{staff.get('name', '')} {staff.get('surname', '')}")
        
        # Mesaili BQ varsa onu kullan, yoksa dashboard BQ'yu kullan
        mesaili_bq = staff.get("mesaili_bq")
        dashboard_bq = _to_float(_safe_get(dash, "user_bq_point_avg", "bq_score", "bq"))
        bq = mesaili_bq if mesaili_bq is not None else dashboard_bq

        productive_sec = _safe_get(dash, "productive_duration_avg")
        distractive_sec = _safe_get(dash, "distractive_duration_avg")
        prod_dist_total = None
        total_productive_pct = None
        if productive_sec is not None and distractive_sec is not None:
            prod_dist_total = _seconds_to_hhmm(productive_sec + distractive_sec)
            expected_sec = _safe_get(dash, "expected_work_duration_avg")
            if expected_sec and expected_sec > 0:
                total_productive_pct = round((productive_sec / expected_sec) * 100, 1)

        values = [
            name,
            staff.get("team_name", "-"),
            _calc_work_days(dash),
            _calc_worked_days(dash),
            bq,
            _format_time(_safe_get(dash, "all_start_time", "start_time_avg")),
            _format_time(_safe_get(dash, "all_end_time", "end_time_avg")),
            _seconds_to_hhmm(_safe_get(dash, "expected_work_duration_avg")),
            _seconds_to_hhmm(productive_sec),
            _seconds_to_hhmm(distractive_sec),
            prod_dist_total,
            total_productive_pct,
            _seconds_to_hhmm(_safe_get(dash, "idle_duration_avg")),
            _bytes_to_gb(_safe_get(dash, "all_download_bytes", "download_total")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # BQ hücresine renk uygula (5. kolon)
        bq_cell = ws.cell(row=row_idx, column=5)
        if bq is not None:
            bq_cell.font = bq_low_font if bq < 100 else bq_ok_font

    # Sütun genişlikleri
    col_widths = [25, 30, 14, 12, 10, 12, 12, 14, 12, 12, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # Kaydet
    os.makedirs(output_dir, exist_ok=True)
    safe_name = team_name.replace(" ", "_").replace("/", "-")
    filename = f"{safe_name}_{report_date}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ Excel oluşturuldu: {filepath}")
    return filepath


# ============================================================
# ANA RAPOR FONKSİYONU
# ============================================================

def generate_full_gmy_report(date: str, output_dir: str = "generated_reports") -> dict:
    """
    Tam aylık GMY raporu oluştur.

    1. Berqun'dan staff listesini çek
    2. Team_name'e göre GMY'ye grupla
    3. Her personelin dashboard verisini çek
    4. BQ < 100 ve VPN listelerini oluştur
    5. HTML mail + Excel dosyaları oluştur

    Returns:
        {
            "gmy_reports": {
                "Teknoloji GMY": {
                    "html": "...",
                    "excel_files": ["path1.xlsx", ...],
                    "bq_below_100": [...],
                    "vpn_missing": [...],
                    "total_staff": 50,
                    "teams": {...}
                }
            }
        }
    """
    client = get_berqun_client()
    staff_list = client.get_staff_list()

    # GMY bazında grupla
    gmy_groups = group_staff_by_gmy(staff_list)

    result = {
        "date": date,
        "generated_at": datetime.now().isoformat(),
        "gmy_reports": {}
    }

    for gmy_name, gmy_data in gmy_groups.items():
        print(f"\n📊 {gmy_name} raporu hazırlanıyor...")
        gmy_report = {
            "total_staff": gmy_data["total_staff"],
            "teams": {},
            "bq_below_100": [],
            "vpn_missing": [],
            "excel_files": [],
            "html": ""
        }

        # Her takım için dashboard verilerini çek
        for team_name, team_staff in gmy_data["teams"].items():
            enriched_staff = []
            for staff in team_staff:
                guid = staff.get("guid") or staff.get("staff_guid")
                if guid:
                    try:
                        dashboard = client.get_user_dashboard(guid, date, PERIOD_MONTH)
                        staff_copy = dict(staff)
                        staff_copy["dashboard"] = dashboard
                        bq = _to_float(_safe_get(dashboard, "user_bq_point_avg"))
                        staff_copy["bq_score"] = bq

                        # BQ < 100 kontrol
                        if bq is not None and bq < 100:
                            gmy_report["bq_below_100"].append({
                                "name": staff.get("full_name", f"{staff.get('name', '')} {staff.get('surname', '')}"),
                                "team_name": team_name,
                                "bq_score": bq,
                                "email": staff.get("email", ""),
                                "staff_guid": guid
                            })

                        enriched_staff.append(staff_copy)
                    except Exception as e:
                        print(f"  ⚠️ {staff.get('full_name', guid)}: {e}")
                        enriched_staff.append(staff)

            gmy_report["teams"][team_name] = enriched_staff

            # Excel oluştur
            excel_path = generate_team_excel(team_name, enriched_staff, date, output_dir)
            if excel_path:
                gmy_report["excel_files"].append(excel_path)

        # VPN kontrolü
        gmy_report["vpn_missing"] = client._check_vpn_compliance(date)

        # Rapor tarih formatı
        try:
            dt = datetime.strptime(date, "%Y-%m-%d")
            months_tr = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                        "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
            report_date_str = f"{months_tr[dt.month]} {dt.year}"
        except:
            report_date_str = date

        # HTML mail oluştur
        gmy_report["html"] = generate_mail_html(
            gmy_name=gmy_name,
            report_date=report_date_str,
            teams_data=gmy_report["teams"],
            bq_below_100=gmy_report["bq_below_100"],
            vpn_missing=gmy_report["vpn_missing"]
        )

        # HTML dosyaya kaydet
        os.makedirs(output_dir, exist_ok=True)
        safe_gmy = gmy_name.replace(" ", "_").replace("/", "-")
        html_path = os.path.join(output_dir, f"{safe_gmy}_{date}_mail.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(gmy_report["html"])
        gmy_report["html_file"] = html_path

        result["gmy_reports"][gmy_name] = gmy_report
        print(f"  ✅ {gmy_name}: {gmy_data['total_staff']} personel, "
              f"BQ<100: {len(gmy_report['bq_below_100'])}, "
              f"VPN missing: {len(gmy_report['vpn_missing'])}")

    return result


# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================

def _safe_get(data: dict, *keys):
    """Dict'den birden fazla olası key ile değer al"""
    if not isinstance(data, dict):
        return None
    for key in keys:
        if key in data:
            return data[key]
    # İç içe arama
    for nested in ["summary", "data", "result"]:
        if nested in data and isinstance(data[nested], dict):
            result = _safe_get(data[nested], *keys)
            if result is not None:
                return result
    return None


def _to_float(val) -> Optional[float]:
    """Değeri float'a çevir"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _fmt(val) -> str:
    """Değeri görüntüleme formatına çevir"""
    if val is None:
        return "-"
    if isinstance(val, float):
        return f"{val:.1f}" if val != int(val) else str(int(val))
    return str(val)


def _seconds_to_hhmm(seconds) -> Optional[str]:
    """Saniyeyi HH:MM formatına çevir"""
    if seconds is None:
        return None
    try:
        total_sec = float(seconds)
        hours = int(total_sec // 3600)
        minutes = int((total_sec % 3600) // 60)
        return f"{hours:02d}:{minutes:02d}"
    except (ValueError, TypeError):
        return None


def _bytes_to_gb(bytes_val) -> Optional[str]:
    """Byte'ı GB olarak formatla"""
    if bytes_val is None:
        return None
    try:
        gb = float(bytes_val) / (1024 ** 3)
        return f"{gb:.2f}"
    except (ValueError, TypeError):
        return None


def _calc_work_days(dash: dict) -> Optional[int]:
    """Dashboard activities listesinden çalışma günü sayısını hesapla"""
    if not isinstance(dash, dict):
        return None
    # expected_work_duration_total / expected_work_duration_avg = iş günü
    total = _safe_get(dash, "expected_work_duration_total")
    avg = _safe_get(dash, "expected_work_duration_avg")
    if total and avg and avg > 0:
        return round(total / avg)
    # Alternatif: activities listesi (tarihlere bak)
    activities = dash.get("activities", [])
    if activities:
        dates = set()
        for act in activities:
            name = act.get("name", "")
            if len(name) == 10 and name[4] == "-":  # YYYY-MM-DD format
                dates.add(name)
        if dates:
            return len(dates)
    return None


def _calc_worked_days(dash: dict) -> Optional[int]:
    """Çalışılan gün sayısını hesapla"""
    if not isinstance(dash, dict):
        return None
    total = _safe_get(dash, "work_duration_total")
    avg = _safe_get(dash, "work_duration_avg")
    if total and avg and avg > 0:
        return round(total / avg)
    return None
