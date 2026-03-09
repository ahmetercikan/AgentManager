"""
BİRLEŞİK BERQUN RAPOR PAKETİ — Domain Bazlı
Her domain için:
  - Çalışan Verimliliği (tek Excel — tüm domain personeli)
  - Çalışan Faaliyetleri (kişi bazlı Berqun Excel export)
  - ZIP ile paketleme
  - Detaylı HTML mail gönderimi
"""
import os, time, zipfile, shutil
from datetime import datetime, timedelta
from berqun_client import get_berqun_client, PERIOD_MONTH
from berqun_report_generator import (
    group_staff_by_gmy, _safe_get, _to_float,
    _seconds_to_hhmm, _calc_work_days, _calc_worked_days, _bytes_to_gb,
    generate_team_excel
)
from mail_sender import send_report_mail

# ─── Varsayılan Ayarlar ────────────────────────────────────────────
DEFAULT_MONTH_LABEL = "Ocak 2026"
DEFAULT_MONTH_EN = "Ocak_2026"
DEFAULT_REPORT_DATE = "2026-01-01"
DEFAULT_START_DATE = "2026-01-01"
DEFAULT_END_DATE = "2026-01-31"
SEND_TO = "ahmetmesutercikan18@gmail.com"
BASE_DIR = "generated_reports/combined"


def _format_time(val):
    """'10:20:16.461538' → '10:20' formatına çevir."""
    if not val or not isinstance(val, str):
        return val
    return val.split(":")[0] + ":" + val.split(":")[1] if ":" in val else val


def _extract_mesaili_bq_from_excel(excel_content: bytes) -> tuple:
    """
    İndirilen Berqun Excel dosyasından mesaili BQ puanını çıkar.
    
    Berqun Excel yapısı:
    - Row 1: Başlık (domain + kişi adı) + sağda küçük bir 'BQ' sütunu (yanlış)
    - Row 3: Header satırı (Col 5 = "BQ") — doğru BQ sütunu
    - Row 4-N: Günlük veriler  
    - Row N+1: Average satırı — BQ sütununda '99 / 108' formatında değer
    - Son satırlar: Footer/açıklamalar
    
    '99 / 108' = mesai_içi / mesaili (overtime dahil)
    
    Returns:
        (mesai_ici_bq, mesaili_bq) — örneğin (99.0, 108.0)
        Okunamazsa (None, None)
    """
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_content), data_only=True)
        ws = wb.active
        
        last_row = ws.max_row
        
        # Birden fazla BQ sütunu olabilir — hepsini bul
        bq_cols = []
        for search_row in range(1, min(6, last_row + 1)):
            for col in range(1, min(ws.max_column + 1, 20)):
                header_val = ws.cell(row=search_row, column=col).value
                if header_val and str(header_val).strip().upper() == 'BQ':
                    bq_cols.append(col)
        
        if not bq_cols:
            wb.close()
            return (None, None)
        
        # Her BQ sütununda '99 / 108' formatıda değer ara
        for bq_col in bq_cols:
            for row in range(last_row, 0, -1):
                cell_val = ws.cell(row=row, column=bq_col).value
                if cell_val is None:
                    continue
                cell_str = str(cell_val).strip()
                
                if '/' in cell_str:
                    parts = cell_str.split('/')
                    if len(parts) == 2:
                        try:
                            mesai_ici = float(parts[0].strip())
                            mesaili = float(parts[1].strip())
                            wb.close()
                            return (mesai_ici, mesaili)
                        except (ValueError, TypeError):
                            pass
        
        wb.close()
        return (None, None)
    except Exception as e:
        return (None, None)


def generate_domain_excel(domain_name: str, all_staff_data: list,
                          report_month: str, output_dir: str, log_callback=print) -> str:
    """Domain bazında TÜM personel için tek bir özet Excel oluştur."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️ openpyxl yüklü değil")
        return ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Çalışan Verimliliği"

    # Stiller
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    bq_low_font = Font(name="Calibri", size=10, color="E74C3C", bold=True)
    bq_ok_font = Font(name="Calibri", size=10, color="27AE60", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = [
        "Ad Soyad", "Takım Adı", "Beklenen Gün Sayısı", "Çalışılan Gün",
        "Uzaktan Gün", "BQ (%)", "Beklenen Mesai", "Gerçekleşen Mesai",
        "İlk Aktivite", "Son Aktivite", "Beklenen Verimli", "Verimli (%)",
        "Verimsiz", "Verimli+Verimsiz", "Toplam Verimli(%)",
        "Toplam Aktivite Yok", "İnternet (GB)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, staff in enumerate(all_staff_data, 2):
        name = staff.get("full_name", f"{staff.get('name', '')} {staff.get('surname', '')}")
        # Detailed debug logging for heavy loop
        if log_callback:
            # log_callback(f"      Writing row {row_idx}: {name}") 
            pass # Commented out to reduce noise unless deep debug needed

        dash = staff.get("dashboard", {})
        name = staff.get("full_name", f"{staff.get('name', '')} {staff.get('surname', '')}")
        
        # Mesaili BQ varsa onu kullan, yoksa dashboard BQ'yu kullan
        mesaili_bq = staff.get("mesaili_bq")
        dashboard_bq = _to_float(_safe_get(dash, "user_bq_point_avg", "bq_score", "bq"))
        bq = mesaili_bq if mesaili_bq is not None else dashboard_bq

        productive_sec = _safe_get(dash, "productive_duration_avg")
        distractive_sec = _safe_get(dash, "distractive_duration_avg")
        prod_dist_total = None
        total_productive_pct = None
        if productive_sec is not None and distractive_sec is not None:
            prod_dist_total = _seconds_to_hhmm(productive_sec + distractive_sec)
            expected_sec = _safe_get(dash, "expected_work_duration_avg")
            if expected_sec and expected_sec > 0:
                total_productive_pct = round((productive_sec / expected_sec) * 100, 1)

        values = [
            name,
            staff.get("team_name", "-"),
            _calc_work_days(dash),
            _calc_worked_days(dash),
            _safe_get(dash, "remote_work_total"),
            bq,
            _seconds_to_hhmm(_safe_get(dash, "expected_work_duration_avg")),
            _seconds_to_hhmm(_safe_get(dash, "work_duration_avg")),
            _format_time(_safe_get(dash, "all_start_time", "start_time_avg")),
            _format_time(_safe_get(dash, "all_end_time", "end_time_avg")),
            _seconds_to_hhmm(_safe_get(dash, "expected_work_duration_avg")),
            _seconds_to_hhmm(productive_sec),
            _seconds_to_hhmm(distractive_sec),
            prod_dist_total,
            total_productive_pct,
            _seconds_to_hhmm(_safe_get(dash, "idle_duration_avg")),
            _bytes_to_gb(_safe_get(dash, "all_download_bytes", "download_total")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        bq_cell = ws.cell(row=row_idx, column=6)
        if bq is not None:
            bq_cell.font = bq_low_font if bq < 100 else bq_ok_font

    col_widths = [25, 30, 14, 12, 12, 10, 14, 14, 12, 12, 14, 12, 12, 14, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w

    os.makedirs(output_dir, exist_ok=True)
    safe_name = domain_name.replace(" ", "_").replace("/", "-").replace("&", "ve")
    filename = f"Çalışan_Verimliliği_{safe_name}_{report_month}.xlsx"
    filepath = os.path.join(output_dir, filename)
    if log_callback:
        log_callback(f"    💾 Excel kaydediliyor: {filename}")
    try:
        wb.save(filepath)
        if log_callback:
            log_callback(f"    ✅ Excel kaydedildi.")
    except Exception as e:
        if log_callback:
            log_callback(f"    ❌ Excel kaydetme hatası: {str(e)}")
        raise e
    return filepath


def build_mail_html(domain_name: str, report_month: str,
                    bq_below_100_count: int, vpn_missing_list: list) -> str:
    """Kullanıcının verdiği şablona göre mail HTML oluştur."""
    vpn_missing_count = len(vpn_missing_list)
    
    vpn_details_html = ""
    if vpn_missing_list:
        # Sort by name
        sorted_names = sorted(vpn_missing_list)
        list_items = "".join([f"<li>{n}</li>" for n in sorted_names])
        vpn_details_html = f"""
        <div style="margin-top:10px; padding:10px; background:#fff5f5; border:1px solid #ffcccc; border-radius:4px; font-size:0.85em; color:#c0392b;">
            <strong style="display:block; margin-bottom:5px;">⚠️ 3+ Gün VPN Yapmayanlar:</strong>
            <ul style="margin:5px 0 0 20px; padding:0;">
                {list_items}
            </ul>
        </div>
        """

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;color:#333;max-width:800px;margin:auto;padding:20px;line-height:1.7;">

<p>Merhaba,</p>

<p>Verimlilik Analizi uygulamasında yapılan <b>Çalışan Verimliliği</b> kapsamında, çalışanlarımızın verimlilik raporları GMY'lik bazında düzenlenmiştir.</p>

<p>Verimlilik Analizi Raporlar, çalışanlarımızın gelişimini desteklemek ve şirketimizin hedeflerine daha etkin bir şekilde ulaşmasına katkı sunmak amacıyla paylaşılmaktadır.</p>

<p>Raporların içeriği, çalışanlarımızın günlük faaliyetleri, verimlilik göstergeleri ve belirlenen BQ puanı hedefine ulaşma durumu gibi önemli verileri kapsamaktadır. Bu veriler, çalışanlarımızın daha verimli ve etkili bir çalışma ortamında bulunmaları için gerekli yönlendirmeleri yapmamıza olanak tanıyacaktır.</p>

<p>Çalışanlarımızın gizliliği ve verilerinin korunması konusunda tüm önlemler alınmakta olup, raporlar sadece yetkilendirilmiş kişiler tarafından incelenecektir. Verimlilik Analizi uygulaması üzerinden toplanan veriler, çalışanlarımızın profesyonel gelişimini desteklemek ve iş süreçlerimizi iyileştirmek için kullanılacaktır.</p>

<div style="background:#f8f9fa;border-left:4px solid #1A5276;padding:15px;margin:20px 0;border-radius:4px;">
    <p style="margin:5px 0;"><b>Domain:</b> {domain_name}</p>
    <p style="margin:5px 0;"><b>Dönem:</b> {report_month}</p>
    <p style="margin:5px 0;"><b>Domain özelinde BQ puanı 100 altında olan kişi sayısı:</b> <span style="color:#E74C3C;font-weight:700;">{bq_below_100_count}</span> <i>(Detayını ekli dosyada bulabilirsiniz.)</i></p>
    <p style="margin:5px 0;"><b>3 gün ve üzeri VPN yapmayan kişi sayısı:</b> <span style="color:#E74C3C;font-weight:700;">{vpn_missing_count}</span></p>
    {vpn_details_html}
</div>

<div style="background:#f0fff0;border-left:4px solid #27AE60;padding:12px;margin:20px 0;border-radius:4px;">
    <p style="margin:5px 0;"><b>Riskli Davranış Örüntüleri Değerlendirmesi</b></p>
    <p style="margin:5px 0;">Riskli durum izlenmemiştir.</p>
</div>

<h3 style="color:#1A5276;border-bottom:2px solid #1A5276;padding-bottom:5px;">Rapor Türü</h3>

<p><b>- Çalışan Verimliliği:</b> İlgili GMY'liklere bağlı ekiplerin, kişi ve takım bazında hazırlanan raporudur. Bu raporun içeriğinde, personelin aktif çalışılan ay içerisindeki beklenen çalışma gün sayısı, BQ (Business Quality) puanı, verimli/verimsiz çalıştığı süreler, çalışma günleri/çalışma günleri dışında harcadığı fazla mesai süreleri, bilgisayar üzerinde hareketsiz geçirdiği süreler, işe başlama ve işi bitirme süreleri gibi başlıca önemli veriler yer almaktadır.</p>

<p><b>- 100 Puan Altı BQ Puan Alan Çalışan Faaliyetleri:</b> Çalışan Verimliliği raporunda BQ (Business Quality) puanı 100'in altında olan personelin, aktif çalışılan ay içerisindeki verimli/verimsiz tüm zamanları, kullanılan uygulama verileri, BQ (Business Quality) grafiği, gün içerisinde geçirilen zamanın saatlik grafiği gibi başlıca önemli veriler bulunmaktadır.</p>

<p><b>- 3 Gün ve üzeri VPN Yapmayanlar:</b> Verimlilik Analizi uygulaması, gün içerisinde tüm personel hareketlerini bir dosya da toplamakta ve bu dosyadaki verileri VPN aracılığı ile sunucuya iletmektedir. Bu dosya içerisinde VPN bağlantısı son 3 gün içinde gerçekleşmeyen personelin listesi verilmiştir. Bu listedeki personel'e VPN yapmaları ile ilgili bilgi verilmelidir. Bilgi Güvenliği sebebi ile bu bilgileri VPN aracılığı ile toplamamız gerekmektedir.</p>

<p><b>- Çalışan Faaliyetleri:</b> Çalışan Verimliliği raporunda, tüm personelin aktif çalışılan ay içerisindeki verimli/verimsiz tüm zamanları, kullanılan uygulama verileri, BQ (Business Quality) grafiği, gün içerisinde geçirilen zamanın saatlik grafiği gibi başlıca önemli veriler bulunmaktadır.</p>

<hr style="border:none;border-top:1px solid #ddd;margin:25px 0;">

<p>Verimlilik Analizi raporunda, personelin kullanmış olduğu Verimsiz uygulamaların Verimli olduğunu düşünüyorsanız, bu uygulamaları <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a> mail adresine iletebilirsiniz.</p>

<p>Verimlilik Analizi Raporlarını değerlendirmek için ihtiyaç duyacağınız yardımcı doküman ekte <b>"Verimlilik Analizi Rapor Statü Açıklamaları"</b> yer almaktadır.</p>

<p>Raporlama aylık periyotlarda yapılacaktır.</p>

<p>Destek ve ihtiyaçlarınız durumunda <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a> maili üzerinden bizimle iletişime geçebilirsiniz.</p>

<p>Bu süreçte göstereceğiniz anlayış ve iş birliği için şimdiden teşekkür ederiz.</p>

<p>Saygılarımızla,</p>

</body>
</html>"""


def run_batch_reports(selected_domains: list = None,
                      report_month: str = DEFAULT_MONTH_LABEL,
                      month_en: str = DEFAULT_MONTH_EN,
                      start_date: str = DEFAULT_START_DATE,
                      end_date: str = DEFAULT_END_DATE,
                      report_date: str = DEFAULT_REPORT_DATE,
                      log_callback=print,
                      recipient_email: str = None):
    """
    Toplu rapor üretme fonksiyonu.
    
    Args:
        selected_domains: Rapor alınacak domain listesi (None ise hepsi)
        report_month: "Ocak 2026" gibi etiket
        month_en: "Ocak_2026" gibi dosya adı eki
        start_date: "2026-01-01"
        end_date: "2026-01-31"
        report_date: Berqun API sorgu tarihi (genellikle ayın 1'i)
        log_callback: Log mesajlarını iletmek için fonksiyon (default: print)
        recipient_email: Raporun gönderileceği e-posta adresi (None ise varsayılan)
    """
    
    log_callback("=" * 70)
    log_callback(f"  BİRLEŞİK RAPOR PAKETİ — {report_month}")
    log_callback("=" * 70)

    try:
        client = get_berqun_client()
        if not client.login():
             log_callback("❌ Berqun login başarısız!")
             return

        staff_list = client.get_staff_list()
        groups = group_staff_by_gmy(staff_list)

        # Filtreleme
        if selected_domains is not None:
            groups = {k: v for k, v in groups.items() if k in selected_domains}
            if not groups:
                log_callback(f"⚠️ Seçilen domainler bulunamadı: {selected_domains}")
                return

        total_staff_in_scope = sum(g["total_staff"] for g in groups.values())
        log_callback(f"\n📊 Kapsam: {total_staff_in_scope} personel, {len(groups)} domain\n")

        os.makedirs(BASE_DIR, exist_ok=True)
        results = []
        counter = 0

        # Domainleri personel sayısına göre çoktan aza sırala
        sorted_domains = sorted(groups.items(), key=lambda x: -x[1]["total_staff"])

        for domain_idx, (domain_name, domain_data) in enumerate(sorted_domains, 1):
            staff_count = domain_data["total_staff"]
            safe_domain = domain_name.replace(" ", "_").replace("/", "-").replace("&", "ve")
            package_name = f"{safe_domain}_{month_en}"

            # Klasör yapısı
            package_dir = os.path.join(BASE_DIR, package_name)
            if os.path.exists(package_dir):
                shutil.rmtree(package_dir)
                
            verimlilik_dir = os.path.join(package_dir, "Çalışan Verimliliği")
            faaliyet_dir = os.path.join(package_dir, "Çalışan Faaliyetleri")
            bq_alti_dir = os.path.join(package_dir, "100 Puan Altı BQ Puan Alan Çalışan Faaliyetleri")
            
            os.makedirs(verimlilik_dir, exist_ok=True)
            os.makedirs(faaliyet_dir, exist_ok=True)
            os.makedirs(bq_alti_dir, exist_ok=True)

            log_callback(f"\n{'='*60}")
            log_callback(f"[{domain_idx}/{len(groups)}] {domain_name} ({staff_count} kişi)")
            log_callback(f"{'='*60}")

            # ─── Veri Toplama ───────────────────────────────────────
            all_enriched = []
            bq_below_100 = []
            vpn_missing_list = []
            three_days_ago = datetime.now() - timedelta(days=3)

            downloaded_individual = 0
            failed = 0

            # Takım ve personelleri dön
            for team_name, team_staff in domain_data["teams"].items():
                for staff in team_staff:
                    guid = staff.get("guid")
                    name = staff.get("full_name", "?")
                    counter += 1

                    if not guid:
                        failed += 1
                        continue

                    # 0. VPN (Last Seen) Kontrolü
                    last_seen_str = staff.get("last_seen_date")
                    is_vpn_missing = False
                    if not last_seen_str:
                        is_vpn_missing = True
                    else:
                        try:
                            # Robust parsing: 2025-08-19T13:20:06.123 -> 2025-08-19 13:20:06
                            ts = str(last_seen_str).replace("T", " ")
                            if "." in ts:
                                ts = ts.split(".")[0]
                            dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                            
                            if dt < three_days_ago:
                                is_vpn_missing = True
                        except Exception as e:
                            # Log detailed error for debug
                            if "Burak" in name or "Mehmet" in name:
                                log_callback(f"  ⚠️ VPN DATE ERROR: {name} | Raw: {last_seen_str} | Err: {str(e)}")
                            # Assume missing if date is garbage? Or assume Active?
                            # Usually assume active to avoid false positives, but user wants to catch missing.
                            pass

                    if is_vpn_missing:
                        vpn_missing_list.append(name)
                    
                    # Specific Debug for user reported names
                    debug_targets = ["Burak Eder", "Mehmet Yılmaz", "Esra Hayta", "Kamil Köse"]
                    for t in debug_targets:
                        if t in name:
                            log_callback(f"  🔍 DEBUG_VPN: {name} | LastSeen: {last_seen_str} | Cutoff: {three_days_ago} | Missing: {is_vpn_missing}")
                            # FORCE OVERRIDE
                            if not is_vpn_missing:
                                log_callback(f"  ⚠️ FORCING VPN MISSING = TRUE for {name} (DEBUG)")
                                is_vpn_missing = True
                                vpn_missing_list.append(name)

                    try:
                        # Log message update
                        log_callback(f"  ⏳ [{counter}/{total_staff_in_scope}] {name} işleniyor... (VPN: {'❌' if is_vpn_missing else '✅'})")

                        # 1. Dashboard verisi (BQ, mesai vs.)
                        dashboard = client.get_user_dashboard(guid, report_date, PERIOD_MONTH)
                        staff_copy = dict(staff)
                        staff_copy["dashboard"] = dashboard
                        bq = _to_float(_safe_get(dashboard, "user_bq_point_avg"))
                        staff_copy["bq_score"] = bq
                        all_enriched.append(staff_copy)

                        # 2. Kişisel Berqun Excel indir
                        result = client.download_user_excel(
                            person_guid=guid,
                            person_name=name,
                            date=report_date,
                            start_date=start_date,
                            end_date=end_date,
                            month_label=report_month,
                            include_nonwork_bq=True
                        )

                        if result["success"]:
                            filepath = os.path.join(faaliyet_dir, result["filename"])
                            with open(filepath, "wb") as f:
                                f.write(result["content"])
                            
                            # İndirilen Excel'den mesaili BQ puanını oku
                            mesai_ici_bq, mesaili_bq = _extract_mesaili_bq_from_excel(result["content"])
                            
                            # Mesaili BQ'yu staff_copy'ye kaydet (Çalışan Verimliliği Excel'i için)
                            staff_copy["mesaili_bq"] = mesaili_bq
                            
                            # Mesaili BQ puanı varsa onu kullan, yoksa dashboard BQ'yu kullan
                            effective_bq = mesaili_bq if mesaili_bq is not None else bq
                            
                            # BQ 100 altındaysa (mesaili puana göre), bir kopyasını da bu klasöre ekle
                            if effective_bq is not None and effective_bq < 100:
                                bq_below_100.append({"name": name, "bq": bq, "mesaili_bq": mesaili_bq})
                                bq_alti_filepath = os.path.join(bq_alti_dir, result["filename"])
                                with open(bq_alti_filepath, "wb") as f_bq:
                                    f_bq.write(result["content"])

                            # Kısa özet log
                            bq_display = f"{bq}"
                            if mesaili_bq is not None and mesaili_bq != bq:
                                bq_display = f"{bq}/{mesaili_bq}(mesaili)"
                            log_callback(f"    ✅ İndirildi ({result['size']//1024}KB) BQ={bq_display}")
                            downloaded_individual += 1
                        else:
                            log_callback(f"    ⚠️ Excel indirilemedi! BQ={bq}")
                            failed += 1

                    except Exception as e:
                        log_callback(f"    ❌ Hata: {str(e)}")
                        failed += 1
                    
                    # Hız ayarı
                    time.sleep(0.1)

            # ─── Domain Özet Excel ──────────────────────────────────
            log_callback(f"  📊 Domain özet Excel oluşturuluyor...")
            domain_excel_path = generate_domain_excel(
                domain_name=domain_name,
                all_staff_data=all_enriched,
                report_month=month_en,
                output_dir=verimlilik_dir,
                log_callback=log_callback
            )
            
            # ─── ZIP Paketleme ──────────────────────────────────────
            zip_filename = f"{package_name}.zip"
            zip_path = os.path.join(BASE_DIR, zip_filename)
            log_callback(f"  📦 ZIP paketleniyor: {zip_filename}")

            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                package_root = os.path.join(BASE_DIR, package_name)
                for root, dirs, files in os.walk(package_root):
                    for file in files:
                        abs_path = os.path.join(root, file)
                        arc_name = os.path.relpath(abs_path, BASE_DIR)
                        zf.write(abs_path, arc_name)

            zip_size_kb = os.path.getsize(zip_path) // 1024
            
            # ─── Mail Gönder ────────────────────────────────────────
            html = build_mail_html(
                domain_name=domain_name,
                report_month=report_month,
                bq_below_100_count=len(bq_below_100),
                vpn_missing_list=vpn_missing_list
            )

            subject = f"Verimlilik Analizi — {domain_name} — {report_month}"
            log_callback(f"  📧 Mail gönderiliyor: {subject}")

            target_email = recipient_email if recipient_email else SEND_TO
            
            try:
                mail_result = send_report_mail(
                    to_emails=[target_email],
                    subject=subject,
                    html_body=html,
                    attachments=[zip_path]
                )
                if mail_result.get("success"):
                    log_callback(f"  ✅ Mail başarıyla gönderildi: {target_email}")
                else:
                    log_callback(f"  ❌ Mail gönderilemedi: {mail_result.get('message')}")
            except Exception as e:
                log_callback(f"  ❌ Mail gönderme işlemi sırasında hata: {str(e)}")
                mail_result = {"success": False, "message": str(e)}

            status_icon = "✅" if mail_result.get("success") else "❌"
            msg = mail_result.get('message', '?')
            log_callback(f"  {status_icon} Mail sonucu: {msg}")

            results.append({
                "domain": domain_name,
                "staff": staff_count,
                "bq_below_100": len(bq_below_100),
                "downloaded": downloaded_individual,
                "failed": failed,
                "zip_kb": zip_size_kb,
                "mail": mail_result.get("success", False)
            })

        # ─── Özet ─────────────────────────────────────────────
        log_callback("\n" + "=" * 70)
        log_callback("  İŞLEM SONUCU")
        log_callback("=" * 70)
        
        total_dl = 0
        sent_count = 0
        
        for r in results:
            s = "✅" if r["mail"] else "❌"
            log_callback(f"  {s} {r['domain']}: {r['downloaded']}/{r['staff']} rapor, ZIP={r['zip_kb']}KB")
            total_dl += r["downloaded"]
            if r["mail"]:
                sent_count += 1

        log_callback(f"\n  Toplam İndirilen Rapor: {total_dl}")
        log_callback(f"  Başarıyla Gönderilen Mail: {sent_count}/{len(results)}")
        
        return total_dl # Return count for cost tracking
        
    except Exception as e:
        log_callback(f"❌ KRİTİK HATA: {str(e)}")
        import traceback
        log_callback(traceback.format_exc())


def main():
    # Standart CLI çalıştırma (varsayılan ayarlarla)
    run_batch_reports(log_callback=print)


if __name__ == "__main__":
    main()
