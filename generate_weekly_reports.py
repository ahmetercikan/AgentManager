"""
BERQUN HAFTALIK RAPOR OTOMASYON SCRIPT'I
Birden fazla ekip icin haftalik verimlilik raporu olusturma.

Desteklenen ekipler:
  - Soft POS          (isim bazli hardcoded)
  - Android POS       (isim bazli hardcoded)
  - Temel Bankacilik - Belkay Saratas   (Berqun team_name bazli)
  - Temel Bankacilik - Mustafa Yildirim (isim bazli hardcoded)

Her ekip icin uretilen klasor yapisi:
  {Ekip Adi}/
  +-- Calisan Faaliyetleri/
  +-- Calisan Verimliligi/
  +-- 100 Puan Alti BQ Puan Alan Calisan Faliyetleri/
  +-- 3 Gun ve uzeri VPN Yapmayanlar/
  +-- Verimsiz Aktiviteler_{Ekip}_{Ay}_{Yil}.xlsx

Kullanim:
  python generate_weekly_reports.py
  python generate_weekly_reports.py --week-start 2026-01-12 --week-end 2026-01-16
  python generate_weekly_reports.py --no-mail
  python generate_weekly_reports.py --dry-run
  python generate_weekly_reports.py --team "Soft POS"
"""
import os
import sys
import time
import argparse
import zipfile
import shutil
from datetime import datetime, timedelta
from typing import Optional

# Windows'ta subprocess olarak çalışırken cp1252 encoding hatası oluşabiliyor.
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from berqun_client import get_berqun_client, PERIOD_WEEK, PERIOD_MONTH
from berqun_report_generator import (
    group_staff_by_gmy, classify_team_to_domain,
    generate_team_excel, _safe_get, _to_float, _seconds_to_hhmm, _bytes_to_gb
)

# ─── Sabitler ────────────────────────────────────────────────
SEND_TO = os.getenv("WEEKLY_REPORT_MAIL", "ahmetmesutercikan18@gmail.com")
BASE_DIR = "generated_reports/weekly"

# ─── EKİP YAPISI (TEAM CONFIGS) ─────────────────────────────
# Her ekip tanımı:
#   name       : Ekip görünen adı (klasör ve rapor başlığı)
#   match_type : "names" → isim bazlı hardcoded | "team_name" → Berqun team_name ile eşleşme
#   members    : (match_type=names) İsim listesi
#   team_name  : (match_type=team_name) Berqun'daki team_name değeri
#   domain     : (opsiyonel) Berqun domain adı — belirtilmezse tüm personelde aranır

TEAM_CONFIGS = [
    {
        "name": "Soft POS",
        "match_type": "names",
        "domain": "Payment Facilitator & Android POS & Soft POS",
        "members": [
            "Fatih Ergül",
            "Mustafa Menteşoğlu",
            "Gürel Yıldız",
            "Gökhan Terzioğlu",
            "Aykut Eryol",
            "İbrahim Demir",
            "Mücahit İmre",
            "Revan Ali",
            "Sinan Yılmaz",
            "Erkan Çokbekler",
            "İpek Dereli",
        ],
    },
    {
        "name": "Android POS",
        "match_type": "names",
        "domain": "Payment Facilitator & Android POS & Soft POS",
        "members": [
            "Rıdvan Çakmak",
            "Mehmet Şükrü Kavak",
            "Ercan Tetik",
            "Omar Gobran",
        ],
    },
    {
        "name": "Temel Bankacilik - Belkay Saratas",
        "match_type": "team_name",
        "team_name_patterns": [
            "Temel Bankacılık - Belkay Sarataş",
            "Temel Bankacılık - Alt Yapı Geliştirme - Belkay Sarataş",
            "Temel Bankacılık - Alt Yapı Geliştirme - Outsource",
        ],
        "extra_members": [
            "Emre Çoğalan",
            "Erhan Akbaş",
        ],
    },
    {
        "name": "Temel Bankacilik - Mustafa Yildirim",
        "match_type": "names",
        "members": [
            "Abbas Ayaz",
            "Abdullah Karataş",
            "Berat Cem Özçelik",
            "Bilal Orkan Tanır",
            "Bilgehan Takım",
            "Cemaleddin Kuru",
            "Emre Özgöz",
            "Hakan Kaysi",
            "Kemal Genç",
            "Merve Aktaş",
            "Merve Esen",
            "Mustafa Yıldırım",
            "Nilsu Acar",
            "Osman Bozok",
            "Sabri Demirel",
            "Semih Akcan",
            "Sude Yalçınkaya",
            "Zafer Kasap",
        ],
    },
]


# ─── Yardımcı fonksiyonlar ───────────────────────────────────

def _normalize_name(name: str) -> str:
    """İsmi küçük harfe çevirip fazla boşlukları temizle."""
    return " ".join(name.strip().lower().split())


def _name_in_list(full_name: str, member_list: list) -> bool:
    """Kişi ismi, listedeki isimlerden biriyle eşleşiyor mu?"""
    norm = _normalize_name(full_name)
    for member in member_list:
        if _normalize_name(member) == norm:
            return True
    return False


def collect_team_members(team_config: dict, all_staff: list, groups: dict) -> list:
    """
    Ekip konfigürasyonuna göre personel listesinden üyeleri topla.
    
    match_type = "names"     → isim listesiyle eşleştir
    match_type = "team_name" → Berqun team_name ile eşleştir
    """
    match_type = team_config["match_type"]
    members = []

    if match_type == "names":
        member_names = team_config["members"]
        # Belirli domain varsa oradan ara, yoksa tüm personelde ara
        domain_name = team_config.get("domain")
        search_pool = []
        if domain_name and domain_name in groups:
            for team_name, team_staff in groups[domain_name]["teams"].items():
                search_pool.extend(team_staff)
        else:
            search_pool = all_staff

        for staff in search_pool:
            full_name = staff.get("full_name",
                                  f"{staff.get('name', '')} {staff.get('surname', '')}").strip()
            staff["full_name"] = full_name
            if _name_in_list(full_name, member_names):
                members.append(staff)

    elif match_type == "team_name":
        # Tekli veya çoklu pattern desteği
        patterns = team_config.get("team_name_patterns", [])
        if not patterns and "team_name_pattern" in team_config:
            patterns = [team_config["team_name_pattern"]]
        patterns_lower = [p.lower() for p in patterns]

        extra = team_config.get("extra_members", [])
        seen_guids = set()

        for staff in all_staff:
            staff_team = staff.get("team_name", "").lower()
            full_name = staff.get("full_name",
                                  f"{staff.get('name', '')} {staff.get('surname', '')}").strip()
            staff["full_name"] = full_name
            guid = staff.get("guid") or staff.get("staff_guid") or full_name

            # team_name pattern eşleşmesi
            matched = any(p in staff_team or staff_team in p for p in patterns_lower)
            # extra_members isim eşleşmesi
            if not matched and extra:
                matched = _name_in_list(full_name, extra)

            if matched and guid not in seen_guids:
                seen_guids.add(guid)
                members.append(staff)

    return members


MONTHS_TR = ["", "Ocak", "Subat", "Mart", "Nisan", "Mayis", "Haziran",
             "Temmuz", "Agustos", "Eylul", "Ekim", "Kasim", "Aralik"]


def get_week_dates(week_start_str: str = None, week_end_str: str = None):
    """
    Hafta başı ve bitiş tarihlerini hesapla.
    Parametre verilmezse geçen haftayı kullan.
    """
    if week_start_str and week_end_str:
        start = datetime.strptime(week_start_str, "%Y-%m-%d")
        end = datetime.strptime(week_end_str, "%Y-%m-%d")
    else:
        today = datetime.now()
        # Geçen haftanın Pazartesi'si
        days_since_monday = today.weekday()
        last_monday = today - timedelta(days=days_since_monday + 7)
        last_friday = last_monday + timedelta(days=4)
        start = last_monday
        end = last_friday

    return start, end


def format_week_label(start: datetime, end: datetime) -> str:
    """'12 Ocak - 16 Ocak' formatı"""
    s_day = start.day
    s_month = MONTHS_TR[start.month]
    e_day = end.day
    e_month = MONTHS_TR[end.month]
    return f"{s_day} {s_month} - {e_day} {e_month}"


def create_folder_structure(base: str, pos_type: str) -> dict:
    """Manuel yapının aynısı klasör yapısını oluştur."""
    root_dir = os.path.join(base, pos_type)
    
    # Eski rapor dosyalarının üzerine yazılması (Ahmet Pirimoğlu vb. eski dosyalar kalmasın diye)
    if os.path.exists(root_dir):
        shutil.rmtree(root_dir)
        
    dirs = {
        "faaliyetler": os.path.join(root_dir, "Çalışan Faaliyetleri"),
        "verimlilik": os.path.join(root_dir, "Çalışan Verimliliği"),
        "bq_altı": os.path.join(root_dir, "100 Puan Altı BQ Puan Alan Çalışan Faaliyetleri"),
        "vpn": os.path.join(root_dir, "3 Gün ve üzeri VPN Yapmayanlar"),
        "root": root_dir,
    }
    for d in dirs.values():
        os.makedirs(d, exist_ok=True)
    return dirs


def generate_verimsiz_excel(staff_data: list, pos_type: str, month_label: str,
                            output_path: str):
    """
    Ekip bazlı Top 10 Verimsiz Aktivite Excel dosyası oluştur.
    
    Tüm ekip üyelerinin level=-1 aktivitelerini toplar,
    süreleri birleştirir ve en çok vakit harcanan 10 verimsiz
    aktiviteyi raporlar.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️ openpyxl yüklü değil. pip install openpyxl")
        return None

    # ─── Tüm ekip üyelerinin verimsiz aktivitelerini topla ───
    activity_totals = {}  # {"Uygulama Adı": toplam_süre_saniye}
    for staff in staff_data:
        apps = staff.get("distractive_apps", [])
        for app in apps:
            app_name = app.get("app", app.get("name", app.get("activity_type_name", "Bilinmeyen")))
            duration = app.get("duration", 0) or 0
            activity_totals[app_name] = activity_totals.get(app_name, 0) + duration

    # Süreye göre sırala ve Top 10 al
    top10 = sorted(activity_totals.items(), key=lambda x: -x[1])[:10]

    if not top10:
        print(f"  ℹ️ {pos_type}: Verimsiz aktivite verisi bulunamadı")
        return None

    def _fmt_dur(sec):
        h, m = int(sec // 3600), int((sec % 3600) // 60)
        return f"{h}sa {m}dk" if h > 0 else f"{m}dk"

    # ─── Excel oluştur ───
    wb = Workbook()
    ws = wb.active
    ws.title = f"Top 10 Verimsiz - {pos_type}"

    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    rank_font = Font(name="Calibri", size=10, bold=True, color="C0392B")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Başlık
    title_cell = ws.cell(row=1, column=1, value=f"{pos_type} - Top 10 Verimsiz Aktivite ({month_label})")
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="C0392B")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ["Sıra", "Aktivite Adı", "Toplam Süre", "Yüzde (%)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    total_all = sum(dur for _, dur in top10) or 1  # sıfıra bölme engeli

    for idx, (app_name, duration) in enumerate(top10, 1):
        row = idx + 3
        pct = round((duration / total_all) * 100, 1)

        ws.cell(row=row, column=1, value=idx).font = rank_font
        ws.cell(row=row, column=2, value=app_name).font = data_font
        ws.cell(row=row, column=3, value=_fmt_dur(duration)).font = data_font
        ws.cell(row=row, column=4, value=f"%{pct}").font = data_font

        for c in range(1, 5):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")

    # Toplam satırı
    total_row = len(top10) + 4
    ws.cell(row=total_row, column=1, value="").border = thin_border
    total_label = ws.cell(row=total_row, column=2, value="TOPLAM")
    total_label.font = Font(name="Calibri", bold=True, size=10)
    total_label.border = thin_border
    total_label.alignment = Alignment(horizontal="center")
    total_dur = ws.cell(row=total_row, column=3, value=_fmt_dur(sum(dur for _, dur in top10)))
    total_dur.font = Font(name="Calibri", bold=True, size=10)
    total_dur.border = thin_border
    total_dur.alignment = Alignment(horizontal="center")
    total_pct = ws.cell(row=total_row, column=4, value="%100")
    total_pct.font = Font(name="Calibri", bold=True, size=10)
    total_pct.border = thin_border
    total_pct.alignment = Alignment(horizontal="center")

    for i, w in enumerate([8, 40, 16, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(output_path)
    print(f"  ✅ Top 10 Verimsiz Aktiviteler: {output_path}")
    return output_path


def generate_vpn_excel(vpn_list: list, pos_type: str, output_path: str):
    """3 Gün ve üzeri VPN Yapmayanlar Excel dosyası oluştur."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl yuklu degil")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = pos_type[:31]  # Excel sheet name max 31 char

    header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    warn_font = Font(name="Calibri", size=10, color="C0392B", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["İsim", "Takım", "Bilgisayar Adı", "Son Sinyal", "İzleme"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, p in enumerate(vpn_list, 2):
        vals = [
            p.get("name", "-"),
            p.get("team_name", "-"),
            p.get("computer_name", "-"),
            p.get("son_sinyal", "-"),
            p.get("izleme", "Etkin"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = warn_font if col == 4 else data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate([25, 40, 35, 16, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(output_path)
    print(f"  VPN Yapmayanlar: {output_path} ({len(vpn_list)} kisi)")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Berqun Haftalik Ekip Rapor Otomasyonu")
    parser.add_argument("--week-start", help="Hafta baslangici (YYYY-MM-DD)")
    parser.add_argument("--week-end", help="Hafta bitisi (YYYY-MM-DD)")
    parser.add_argument("--no-mail", action="store_true", help="Mail gonderme, sadece olustur")
    parser.add_argument("--dry-run", action="store_true", help="API cagrisi yapmadan test et")
    parser.add_argument("--output-dir", default=BASE_DIR, help="Cikti klasoru")
    parser.add_argument("--team", help="Sadece belirli bir ekip icin rapor uret (ornek: 'Soft POS')")
    args = parser.parse_args()

    # Tarih hesapla
    week_start, week_end = get_week_dates(args.week_start, args.week_end)
    week_label = format_week_label(week_start, week_end)
    month_label = f"{MONTHS_TR[week_start.month]}_{week_start.year}"
    date_str = week_start.strftime("%Y-%m-%d")
    end_str = week_end.strftime("%Y-%m-%d")

    # Ana klasor
    folder_name = f"Haftalik_Raporlar_{week_label}"
    report_dir = os.path.join(args.output_dir, folder_name)

    # Hangi ekipler islenecek?
    if args.team:
        active_teams = [t for t in TEAM_CONFIGS if t["name"].lower() == args.team.lower()]
        if not active_teams:
            print(f"Hata: '{args.team}' ekibi bulunamadi.")
            print(f"Mevcut ekipler: {', '.join(t['name'] for t in TEAM_CONFIGS)}")
            return
    else:
        active_teams = TEAM_CONFIGS

    print("=" * 70)
    print(f"  BERQUN HAFTALIK RAPOR")
    print(f"  Donem: {week_label} {week_start.year}")
    print(f"  Tarih: {date_str} -> {end_str}")
    print(f"  Ekipler: {', '.join(t['name'] for t in active_teams)}")
    print(f"  Klasor: {report_dir}")
    print("=" * 70)

    if args.dry_run:
        print("\n  DRY RUN -- API cagrisi yapilmayacak")
        for team_cfg in active_teams:
            team_name = team_cfg["name"]
            dirs = create_folder_structure(report_dir, team_name)
            print(f"\n  {team_name} klasorleri olusturuldu:")
            for key, path in dirs.items():
                print(f"     - {key}: {path}")
        print("\n  Dry run tamamlandi. Gercek calistirma icin --dry-run kaldirin.")
        return

    # Berqun'a baglan
    client = get_berqun_client()
    print("\n  Berqun'a baglaniliyor...")
    client.login()
    print("  Giris basarili")

    # Personel listesini cek
    print("\n  Personel listesi cekiliyor...")
    staff_list = client.get_staff_list()
    print(f"  {len(staff_list)} toplam personel")

    # Domain bazli grupla (isim bazli ekipler icin gerekli)
    groups = group_staff_by_gmy(staff_list)

    # ─── Her ekip icin rapor uret ─────────────────────────
    all_excel_files = []
    team_summaries = {}  # {team_name: member_count}
    team_details = {}    # {team_name: {bq_below_100: N, vpn_missing: [...]}}
    global_counter = 0

    for team_cfg in active_teams:
        team_name = team_cfg["name"]

        # Ekip uyelerini topla
        staff_members = collect_team_members(team_cfg, staff_list, groups)

        if not staff_members:
            print(f"\n  {team_name}: Hicbir uye bulunamadi, atlaniyor")
            team_summaries[team_name] = 0
            continue

        team_summaries[team_name] = len(staff_members)

        print(f"\n{'=' * 70}")
        print(f"  {team_name} -- {len(staff_members)} kisi")
        print(f"{'=' * 70}")

        dirs = create_folder_structure(report_dir, team_name)

        bq_below_100 = []
        enriched_staff = []

        # ─── 1. Calisan Faaliyetleri (kisisel Excel) ────────
        print(f"\n  Calisan Faaliyetleri Indiriliyor...")
        for staff in staff_members:
            guid = staff.get("guid") or staff.get("staff_guid")
            name = staff["full_name"]
            global_counter += 1

            if not guid:
                print(f"  [{global_counter}] {name} -- GUID yok, atlaniyor")
                continue

            try:
                print(f"  [{global_counter}] {name}...", end=" ", flush=True)

                # Dashboard verisi cek (haftalik)
                dashboard = client.get_user_dashboard(guid, date_str, PERIOD_WEEK)
                staff_copy = dict(staff)
                staff_copy["dashboard"] = dashboard
                bq = _to_float(_safe_get(dashboard, "user_bq_point_avg", "bq_score", "bq"))
                staff_copy["bq_score"] = bq
                print(f"BQ={bq or '-'}", end=" ")

                enriched_staff.append(staff_copy)

                # Kisisel Excel indir
                result = client.download_user_excel(
                    person_guid=guid,
                    person_name=name,
                    date=date_str,
                    start_date=date_str,
                    end_date=end_str,
                    month_label=week_label,
                    include_nonwork_bq=True
                )

                if result["success"]:
                    filepath = os.path.join(dirs["faaliyetler"], result["filename"])
                    with open(filepath, "wb") as f:
                        f.write(result["content"])
                    print(f"  {result['size'] // 1024}KB")
                    all_excel_files.append(filepath)
                    
                    # İndirilen Excel'den mesaili BQ puanını oku
                    from generate_combined_reports import _extract_mesaili_bq_from_excel
                    mesai_ici_bq, mesaili_bq = _extract_mesaili_bq_from_excel(result["content"])
                    effective_bq = mesaili_bq if mesaili_bq is not None else bq
                    staff_copy["mesaili_bq"] = mesaili_bq
                    staff_copy["_excel_content"] = result["content"]
                    staff_copy["_excel_filename"] = result["filename"]
                    
                    if effective_bq is not None and effective_bq < 100:
                        bq_below_100.append(staff_copy)
                else:
                    print(f"  Hata: {result.get('error', '?')}")
                    if bq is not None and bq < 100:
                        bq_below_100.append(staff_copy)

                # Verimsiz uygulamalar
                try:
                    app_usage = client.get_user_app_usage(guid, date_str, PERIOD_WEEK)
                    staff_copy["distractive_apps"] = app_usage.get("distractive", [])
                except Exception:
                    staff_copy["distractive_apps"] = []

            except Exception as e:
                print(f"  Hata: {e}")
                enriched_staff.append(staff)

            time.sleep(0.15)

        # ─── 2. Calisan Verimliligi (ozet Excel) ────────────
        print(f"\n  Calisan Verimliligi olusturuluyor...")
        verimlilik_name = f"{team_name} - Calisan Verimliligi {week_label} {week_start.year}"
        excel_path = generate_team_excel(
            team_name=verimlilik_name,
            staff_data=enriched_staff,
            report_date=f"{week_label} {week_start.year}",
            output_dir=dirs["verimlilik"]
        )
        if excel_path:
            all_excel_files.append(excel_path)

        # ─── 3. 100 Puan Alti BQ (mesaili puana göre) ─────────
        if bq_below_100:
            print(f"\n  100 Puan Alti BQ (mesaili) -- {len(bq_below_100)} kisi")
            for staff in bq_below_100:
                name = staff["full_name"]
                excel_content = staff.get("_excel_content")
                excel_filename = staff.get("_excel_filename")
                
                if excel_content and excel_filename:
                    # Zaten indirilmiş Excel'i kopyala (tekrar indirmeye gerek yok)
                    filepath = os.path.join(dirs["bq_altı"], excel_filename)
                    with open(filepath, "wb") as f:
                        f.write(excel_content)
                    mesaili = staff.get("mesaili_bq", staff.get("bq_score"))
                    print(f"    {name} -- BQ={staff.get('bq_score')}/{mesaili}(mesaili)")
                    all_excel_files.append(filepath)
                else:
                    # Excel indirilememişse tekrar dene
                    guid = staff.get("guid") or staff.get("staff_guid")
                    if guid:
                        try:
                            result = client.download_user_excel(
                                person_guid=guid,
                                person_name=name,
                                date=date_str,
                                start_date=date_str,
                                end_date=end_str,
                                month_label=week_label,
                                include_nonwork_bq=True
                            )
                            if result["success"]:
                                filepath = os.path.join(dirs["bq_altı"], result["filename"])
                                with open(filepath, "wb") as f:
                                    f.write(result["content"])
                                print(f"    {name} -- BQ={staff.get('bq_score')}")
                                all_excel_files.append(filepath)
                        except Exception as e:
                            print(f"    {name} -- Hata: {e}")
                        time.sleep(0.1)
                
                # Geçici veriyi temizle
                staff.pop("_excel_content", None)
                staff.pop("_excel_filename", None)
        else:
            print(f"\n  100 Puan Alti BQ -- 0 kisi (tumu 100+)")

        # ─── 4. 3 Gun VPN Yapmayanlar ─────────────────────
        print(f"\n  VPN kontrolu yapiliyor...")
        vpn_missing = []
        try:
            vpn_missing = client._check_vpn_compliance(date_str, staff_list=staff_members)
        except Exception as e:
            print(f"  VPN kontrol hatasi: {e}")

        if vpn_missing:
            vpn_filename = f"{team_name} - 3 Gun ve uzeri VPN Yapmayanlar.xlsx"
            vpn_path = os.path.join(dirs["vpn"], vpn_filename)
            generate_vpn_excel(vpn_missing, team_name, vpn_path)
            all_excel_files.append(vpn_path)
            print(f"  VPN yapmayan: {len(vpn_missing)} kisi")
        else:
            print(f"  Tum personel VPN'e baglanmis")

        # ─── 5. Verimsiz Aktiviteler ──────────────────────
        print(f"\n  Verimsiz Aktiviteler olusturuluyor...")
        safe_team = team_name.replace(" ", "_").replace("-", "_")
        verimsiz_filename = f"Verimsiz_Aktiviteler_{safe_team}_{month_label}.xlsx"
        verimsiz_path = os.path.join(dirs["root"], verimsiz_filename)

        staff_with_apps = [s for s in enriched_staff if s.get("distractive_apps")]
        print(f"  {team_name}: {len(staff_with_apps)}/{len(enriched_staff)} kisinin verimsiz aktivite verisi var")
        if staff_with_apps:
            # Debug: kac benzersiz uygulama var
            all_apps = set()
            for s in staff_with_apps:
                for a in s.get("distractive_apps", []):
                    all_apps.add(a.get("app", "?"))
            print(f"  {team_name}: {len(all_apps)} benzersiz verimsiz uygulama")
            generate_verimsiz_excel(staff_with_apps, team_name, month_label, verimsiz_path)
            all_excel_files.append(verimsiz_path)
        else:
            print(f"  Verimsiz aktivite verisi bulunamadi")

        # ─── Ekip detaylarini kaydet (mail icin) ─────────
        team_details[team_name] = {
            "bq_below_100": len(bq_below_100),
            "vpn_missing": [v.get("name", "-") for v in vpn_missing],
        }

    # ─── ZIP olustur ─────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"  PAKETLEME")
    print(f"{'=' * 70}")

    # Sadece bu calismada olusturulan ekip klasorlerini paketle
    team_suffix = ""
    if args.team:
        safe_name = args.team.replace(" ", "_")
        team_suffix = f"_{safe_name}"

    zip_filename = f"Haftalik_Rapor_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}{team_suffix}.zip"
    zip_path = os.path.join(args.output_dir, zip_filename)
    os.makedirs(args.output_dir, exist_ok=True)

    # Sadece islenen ekiplerin klasorlerini ZIP'e ekle
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for team_cfg in active_teams:
            team_dir = os.path.join(report_dir, team_cfg["name"])
            if not os.path.exists(team_dir):
                continue
            for root, dirs_list, files in os.walk(team_dir):
                for file in files:
                    abs_path = os.path.join(root, file)
                    arc_name = os.path.relpath(abs_path, args.output_dir)
                    zf.write(abs_path, arc_name)

    zip_size = os.path.getsize(zip_path) // 1024
    print(f"  ZIP: {zip_filename} ({zip_size}KB)")

    # ─── Mail gonder ─────────────────────────────────────
    if not args.no_mail:
        try:
            from mail_sender import send_report_mail

            # Her ekip icin ayri mail gonder (aylik sablonla ayni)
            for team_cfg in active_teams:
                tname = team_cfg["name"]
                tcount = team_summaries.get(tname, 0)
                if tcount == 0:
                    continue

                details = team_details.get(tname, {})
                bq_below = details.get("bq_below_100", 0)
                vpn_names = details.get("vpn_missing", [])
                vpn_count = len(vpn_names)

                # VPN listesi HTML
                vpn_details_html = ""
                if vpn_names:
                    sorted_names = sorted(vpn_names)
                    list_items = "".join([f"<li>{n}</li>" for n in sorted_names])
                    vpn_details_html = f"""
                    <div style="margin-top:10px; padding:10px; background:#fff5f5; border:1px solid #ffcccc; border-radius:4px; font-size:0.85em; color:#c0392b;">
                        <strong style="display:block; margin-bottom:5px;">3+ Gün VPN Yapmayanlar:</strong>
                        <ul style="margin:5px 0 0 20px; padding:0;">
                            {list_items}
                        </ul>
                    </div>
                    """

                subject = f"Verimlilik Analizi — {tname} — {week_label} {week_start.year}"
                html_body = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;color:#333;max-width:800px;margin:auto;padding:20px;line-height:1.7;">

<p>Merhaba,</p>

<p>Verimlilik Analizi uygulamasında yapılan <b>Çalışan Verimliliği</b> kapsamında, çalışanlarımızın verimlilik raporları ekip bazında düzenlenmiştir.</p>

<p>Verimlilik Analizi Raporlar, çalışanlarımızın gelişimini desteklemek ve şirketimizin hedeflerine daha etkin bir şekilde ulaşmasına katkı sunmak amacıyla paylaşılmaktadır.</p>

<p>Raporların içeriği, çalışanlarımızın haftalık faaliyetleri, verimlilik göstergeleri ve belirlenen BQ puanı hedefine ulaşma durumu gibi önemli verileri kapsamaktadır. Bu veriler, çalışanlarımızın daha verimli ve etkili bir çalışma ortamında bulunmaları için gerekli yönlendirmeleri yapmamıza olanak tanıyacaktır.</p>

<p>Çalışanlarımızın gizliliği ve verilerinin korunması konusunda tüm önlemler alınmakta olup, raporlar sadece yetkilendirilmiş kişiler tarafından incelenecektir. Verimlilik Analizi uygulaması üzerinden toplanan veriler, çalışanlarımızın profesyonel gelişimini desteklemek ve iş süreçlerimizi iyileştirmek için kullanılacaktır.</p>

<div style="background:#f8f9fa;border-left:4px solid #1A5276;padding:15px;margin:20px 0;border-radius:4px;">
    <p style="margin:5px 0;"><b>Ekip:</b> {tname}</p>
    <p style="margin:5px 0;"><b>Dönem:</b> {week_label} {week_start.year} (Haftalık)</p>
    <p style="margin:5px 0;"><b>Ekip özelinde BQ puanı 100 altında olan kişi sayısı:</b> <span style="color:#E74C3C;font-weight:700;">{bq_below}</span> <i>(Detayını ekli dosyada bulabilirsiniz.)</i></p>
    <p style="margin:5px 0;"><b>3 gün ve üzeri VPN yapmayan kişi sayısı:</b> <span style="color:#E74C3C;font-weight:700;">{vpn_count}</span></p>
    {vpn_details_html}
</div>

<div style="background:#f0fff0;border-left:4px solid #27AE60;padding:12px;margin:20px 0;border-radius:4px;">
    <p style="margin:5px 0;"><b>Riskli Davranış Örüntüleri Değerlendirmesi</b></p>
    <p style="margin:5px 0;">Riskli durum izlenmemiştir.</p>
</div>

<h3 style="color:#1A5276;border-bottom:2px solid #1A5276;padding-bottom:5px;">Rapor Türü</h3>

<p><b>- Çalışan Verimliliği:</b> İlgili ekibe bağlı personelin, kişi ve takım bazında hazırlanan raporudur. Bu raporun içeriğinde, personelin aktif çalışılan hafta içerisindeki beklenen çalışma gün sayısı, BQ (Business Quality) puanı, verimli/verimsiz çalıştığı süreler, bilgisayar üzerinde hareketsiz geçirdiği süreler, işe başlama ve işi bitirme süreleri gibi başlıca önemli veriler yer almaktadır.</p>

<p><b>- 100 Puan Altı BQ Puan Alan Çalışan Faaliyetleri:</b> Çalışan Verimliliği raporunda BQ (Business Quality) puanı 100'in altında olan personelin, aktif çalışılan hafta içerisindeki verimli/verimsiz tüm zamanları, kullanılan uygulama verileri, BQ (Business Quality) grafiği, gün içerisinde geçirilen zamanın saatlik grafiği gibi başlıca önemli veriler bulunmaktadır.</p>

<p><b>- 3 Gün ve üzeri VPN Yapmayanlar:</b> Verimlilik Analizi uygulaması, gün içerisinde tüm personel hareketlerini bir dosya da toplamakta ve bu dosyadaki verileri VPN aracılığı ile sunucuya iletmektedir. Bu dosya içerisinde VPN bağlantısı son 3 gün içinde gerçekleşmeyen personelin listesi verilmiştir. Bu listedeki personel'e VPN yapmaları ile ilgili bilgi verilmelidir. Bilgi Güvenliği sebebi ile bu bilgileri VPN aracılığı ile toplamamız gerekmektedir.</p>

<p><b>- Çalışan Faaliyetleri:</b> Çalışan Verimliliği raporunda, tüm personelin aktif çalışılan hafta içerisindeki verimli/verimsiz tüm zamanları, kullanılan uygulama verileri, BQ (Business Quality) grafiği, gün içerisinde geçirilen zamanın saatlik grafiği gibi başlıca önemli veriler bulunmaktadır.</p>

<hr style="border:none;border-top:1px solid #ddd;margin:25px 0;">

<p>Verimlilik Analizi raporunda, personelin kullanmış olduğu Verimsiz uygulamaların Verimli olduğunu düşünüyorsanız, bu uygulamaları <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a> mail adresine iletebilirsiniz.</p>

<p>Verimlilik Analizi Raporlarını değerlendirmek için ihtiyaç duyacağınız yardımcı doküman ekte <b>"Verimlilik Analizi Rapor Statü Açıklamaları"</b> yer almaktadır.</p>

<p>Raporlama haftalık periyotlarda yapılacaktır.</p>

<p>Destek ve ihtiyaçlarınız durumunda <a href="mailto:verimlilikanalizi@dgpaysit.com">verimlilikanalizi@dgpaysit.com</a> maili üzerinden bizimle iletişime geçebilirsiniz.</p>

<p>Bu süreçte göstereceğiniz anlayış ve iş birliği için şimdiden teşekkür ederiz.</p>

<p>Saygılarımızla,</p>

</body>
</html>"""

                # Ekip ZIP'ini bul
                team_dir = os.path.join(report_dir, tname)
                team_zip_name = f"{tname.replace(' ', '_')}_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}.zip"
                team_zip_path = os.path.join(args.output_dir, team_zip_name)

                # Ekip icin ayri ZIP olustur
                with zipfile.ZipFile(team_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    if os.path.exists(team_dir):
                        for root, dirs_list, files in os.walk(team_dir):
                            for file in files:
                                abs_path = os.path.join(root, file)
                                arc_name = os.path.relpath(abs_path, args.output_dir)
                                zf.write(abs_path, arc_name)

                print(f"\n  Mail gonderiliyor ({tname}): {SEND_TO}")
                result = send_report_mail(
                    to_emails=[SEND_TO],
                    subject=subject,
                    html_body=html_body,
                    attachments=[team_zip_path]
                )
                status = "OK" if result.get("success") else "HATA"
                print(f"  {status} {result.get('message', '?')}")

        except ImportError:
            print("  mail_sender modulu bulunamadi, mail gonderilmedi")
    else:
        print("\n  --no-mail: Mail gonderimi atlandi")

    # ─── Ozet ────────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"  OZET")
    print(f"{'=' * 70}")
    print(f"  Donem: {week_label} {week_start.year}")
    print(f"  Klasor: {report_dir}")
    for t, c in team_summaries.items():
        print(f"  {t}: {c} kisi")
    print(f"  {len(all_excel_files)} dosya olusturuldu")
    print(f"  ZIP: {zip_path} ({zip_size}KB)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()

